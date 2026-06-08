import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, date
import json
import io

st.set_page_config(page_title="Controle de Turnos", page_icon="🔧", layout="wide")

PREFIXOS_ALVO = ("GOOC", "GOOH", "GOOL", "GOOK")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ─── Excel download ──────────────────────────────────────────────────────────

def gerar_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos"

    verde_fill = PatternFill("solid", start_color="C6EFCE")
    vermelho_fill = PatternFill("solid", start_color="FFC7CE")
    header_fill = PatternFill("solid", start_color="2E4057")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell_font = Font(name="Arial", size=10)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    ws.row_dimensions[1].height = 22

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        aberto = str(row.get("Turno Aberto", "")) == "✅ Sim"
        fill = verde_fill if aberto else vermelho_fill
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(row[h]))
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda

    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), *(len(str(row[h])) for _, row in df.iterrows()))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Google Sheets Configuration ──────────────────────────────────────────────

def get_gsheet_client():
    creds_dict = st.session_state.get("gsheet_creds")
    if not creds_dict:
        try:
            creds_dict = st.secrets["gcp_service_account"].to_dict()
        except Exception:
            return None
            
    try:
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
            
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        return gspread.authorize(creds)
    except Exception as e:
        st.sidebar.error(f"Erro nas credenciais: {e}")
        return None


def get_spreadsheet_id():
    sid = st.session_state.get("spreadsheet_id", "")
    if not sid:
        try:
            sid = st.secrets["spreadsheet_id"]
        except Exception:
            pass
    return str(sid).strip()


def get_or_create_sheet(client, spreadsheet_id, sheet_name):
    try:
        ss = client.open_by_key(spreadsheet_id)
    except Exception as e:
        st.error(f"Erro ao abrir planilha: {e}")
        return None
    try:
        return ss.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=sheet_name, rows=1000, cols=20)
        return ws


def ensure_headers(ws, headers):
    existing = ws.row_values(1)
    if existing != headers:
        ws.clear()
        ws.append_row(headers)


def load_sheet_df(ws):
    data = ws.get_all_records(default_blank="")
    return pd.DataFrame(data) if data else pd.DataFrame(columns=ws.row_values(1))


# ─── XML / Excel reader ───────────────────────────────────────────────────────

def ler_xml(conteudo: bytes) -> pd.DataFrame:
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError as e:
        raise ValueError(f"XML inválido: {e}")

    NS = "urn:schemas-microsoft-com:office:spreadsheet"

    def tag(nome):
        return f"{{{NS}}}{nome}"

    worksheet = root.find(f".//{tag('Worksheet')}")
    if worksheet is not None:
        table = worksheet.find(tag("Table"))
        if table is None:
            for ws in root.iter(tag("Worksheet")):
                table = ws.find(tag("Table"))
                if table is not None:
                    break
        if table is not None:
            linhas = list(table.findall(tag("Row")))
            if not linhas:
                raise ValueError("SpreadsheetML: nenhuma linha encontrada.")
            ATTR_INDEX = f"{{{NS}}}Index"

            def celulas(row_el):
                vals = []
                idx = 0
                for cell in row_el.findall(tag("Cell")):
                    ss_idx = cell.get(ATTR_INDEX)
                    if ss_idx:
                        alvo = int(ss_idx) - 1
                        while idx < alvo:
                            vals.append("")
                            idx += 1
                    data_el = cell.find(tag("Data"))
                    vals.append((data_el.text or "").strip() if data_el is not None else "")
                    idx += 1
                return vals

            cabecalho = celulas(linhas[0])
            cab_norm = []
            contagem: dict = {}
            for c in cabecalho:
                c = c.strip() if c.strip() else f"_col{len(cab_norm)}"
                contagem[c] = contagem.get(c, 0) + 1
                cab_norm.append(c if contagem[c] == 1 else f"{c}_{contagem[c]}")

            registros = []
            for row_el in linhas[1:]:
                vals = celulas(row_el)
                while len(vals) < len(cab_norm):
                    vals.append("")
                registros.append(dict(zip(cab_norm, vals[: len(cab_norm)])))

            if not registros:
                raise ValueError("SpreadsheetML: tabela sem linhas de dados.")
            df = pd.DataFrame(registros)
            df.columns = [str(c).strip() for c in df.columns]
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)].reset_index(drop=True)
            return df

    filhos = list(root)
    if not filhos:
        raise ValueError("XML sem elementos filhos na raiz.")
    candidatos = filhos
    for _ in range(3):
        netos = []
        for f in candidatos:
            netos.extend(list(f))
        if len(netos) > len(candidatos):
            candidatos = netos
        else:
            break
    amostra = candidatos[0]
    registros = []
    if list(amostra):
        def limpar_tag(t):
            return t.split("}")[-1] if "}" in t else t
        for filho in candidatos:
            registro = {limpar_tag(sub.tag): (sub.text or "").strip() for sub in filho}
            if registro:
                registros.append(registro)
    elif amostra.attrib:
        for filho in candidatos:
            if filho.attrib:
                registros.append(dict(filho.attrib))
    if not registros:
        raise ValueError("Não foi possível extrair registros do XML.")
    df = pd.DataFrame(registros)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def ler_arquivo(arquivo) -> pd.DataFrame:
    nome = arquivo.name.lower()
    conteudo = arquivo.read()
    arquivo.seek(0)
    if nome.endswith(".xml"):
        return ler_xml(conteudo)
    elif nome.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(conteudo))
    else:
        raise ValueError(f"Formato não suportado: {nome}")


def filtrar_e_formatar(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["Prefixo"].astype(str).str.startswith(PREFIXOS_ALVO)
    df = df[mask].copy()

    def fmt_hora(val):
        if pd.isna(val) or str(val).strip() in ("", "NaT"):
            return ""
        try:
            if isinstance(val, str):
                dt = pd.to_datetime(val, errors="coerce")
            else:
                dt = pd.Timestamp(val)
            if pd.isna(dt):
                return str(val)
            return dt.strftime("%H:%M:%S")
        except Exception:
            return str(val)

    result = pd.DataFrame()
    result["Prefixo"] = df["Prefixo"].astype(str)
    result["Início de Turno"] = df["Início de Turno"].apply(fmt_hora)
    result["Previsão de Saída"] = df["Previsão Saída"].apply(fmt_hora)
    result["Intervalo"] = df["Intervalo"].apply(fmt_hora)
    result["Turno Aberto"] = result["Início de Turno"].apply(lambda x: "✅ Sim" if x else "❌ Não")
    result["Data"] = date.today().strftime("%Y-%m-%d")
    return result.reset_index(drop=True)


# ─── Google Sheets: Operações Otimizadas Antiduplicidade ─────────────────────

HEADERS_TURNOS = ["Data", "Prefixo", "Início de Turno", "Previsão de Saída", "Intervalo",
                  "Fim de Intervalo", "Motivo", "Ocorrência", "Turno Aberto"]
HEADERS_HISTORICO = ["Mês", "Semana", "Prefixo", "Turnos no Mês", "Intervalos na Semana",
                     "Intervalos no Mês"]


def salvar_turnos_gs(client, spreadsheet_id, df_dia):
    ws = get_or_create_sheet(client, spreadsheet_id, "Turnos")
    if ws is None:
        return False
    ensure_headers(ws, HEADERS_TURNOS)
    df_atual = load_sheet_df(ws)
    hoje = date.today().strftime("%Y-%m-%d")
    
    # Remove registros do dia atual na memória antes de ressalvar para evitar duplicar na aba Turnos
    if not df_atual.empty and "Data" in df_atual.columns:
        df_atual = df_atual[df_atual["Data"].astype(str) != hoje]
        
    novos = []
    for _, row in df_dia.iterrows():
        novos.append({
            "Data": row["Data"],
            "Prefixo": row["Prefixo"],
            "Início de Turno": row["Início de Turno"],
            "Previsão de Saída": row["Previsão de Saída"],
            "Intervalo": row["Intervalo"],
            "Fim de Intervalo": "",
            "Motivo": "",
            "Ocorrência": "",
            "Turno Aberto": row["Turno Aberto"],
        })
    
    df_final = pd.concat([df_atual, pd.DataFrame(novos)], ignore_index=True)
    ws.clear()
    
    dados_para_enviar = [HEADERS_TURNOS]
    for _, r in df_final.iterrows():
        dados_para_enviar.append([str(r.get(h, "")) for h in HEADERS_TURNOS])
    
    ws.append_rows(dados_para_enviar) 
    return True


def atualizar_historico(client, spreadsheet_id, df_dia):
    ws_turnos = get_or_create_sheet(client, spreadsheet_id, "Turnos")
    ws_hist = get_or_create_sheet(client, spreadsheet_id, "Histórico")
    if ws_turnos is None or ws_hist is None:
        return False
        
    ensure_headers(ws_hist, HEADERS_HISTORICO)
    df_turnos_existentes = load_sheet_df(ws_turnos)
    df_hist = load_sheet_df(ws_hist)
    
    hoje = date.today()
    mes_atual = hoje.strftime("%Y-%m")
    semana_atual = hoje.strftime("%Y-W%W")
    hoje_str = hoje.strftime("%Y-%m-%d")

    for _, row in df_dia.iterrows():
        prefixo = row["Prefixo"]
        abriu = row["Turno Aberto"] == "✅ Sim"
        tem_intervalo = bool(row["Intervalo"])

        # 🔥 PREVENÇÃO DE DUPLICIDADE COORDENADA:
        # Verifica se essa combinação de Prefixo e Data JÁ EXISTIA de fato na planilha física antes deste upload
        ja_existia_hoje = False
        if not df_turnos_existentes.empty and "Data" in df_turnos_existentes.columns:
            match = df_turnos_existentes[
                (df_turnos_existentes["Data"].astype(str) == hoje_str) & 
                (df_turnos_existentes["Prefixo"].astype(str) == prefixo)
            ]
            if not match.empty:
                ja_existia_hoje = True

        # Se já existia, pula o incremento para não adulterar o histórico de forma cumulativa
        if ja_existia_hoje:
            continue

        mask_mes = (df_hist.get("Mês", pd.Series()) == mes_atual) & \
                   (df_hist.get("Prefixo", pd.Series()) == prefixo) & \
                   (df_hist.get("Semana", pd.Series()) == semana_atual)

        if df_hist.empty or not mask_mes.any():
            df_hist = pd.concat([df_hist, pd.DataFrame([{
                "Mês": mes_atual,
                "Semana": semana_atual,
                "Prefixo": prefixo,
                "Turnos no Mês": 1 if abriu else 0,
                "Intervalos na Semana": 1 if tem_intervalo else 0,
                "Intervalos no Mês": 1 if tem_intervalo else 0,
            }])], ignore_index=True)
        else:
            idx = df_hist[mask_mes].index[0]
            if abriu:
                df_hist.at[idx, "Turnos no Mês"] = int(df_hist.at[idx, "Turnos no Mês"] or 0) + 1
            if tem_intervalo:
                df_hist.at[idx, "Intervalos na Semana"] = int(df_hist.at[idx, "Intervalos na Semana"] or 0) + 1
                df_hist.at[idx, "Intervalos no Mês"] = int(df_hist.at[idx, "Intervalos no Mês"] or 0) + 1

    ws_hist.clear()
    dados_historico = [HEADERS_HISTORICO]
    for _, r in df_hist.iterrows():
        dados_historico.append([str(r.get(h, "")) for h in HEADERS_HISTORICO])
        
    ws_hist.append_rows(dados_historico)
    return True


# ─── Páginas do Sistema ───────────────────────────────────────────────────────

def pagina_upload():
    st.title("📋 Controle de Turnos do Dia")
    st.markdown("Faça o upload da base (XML ou Excel) para visualizar e registrar os turnos.")

    arquivo = st.file_uploader("📂 Selecione o arquivo da base", type=["xml", "xlsx", "xls"])

    if arquivo:
        try:
            df_raw = ler_arquivo(arquivo)
            df = filtrar_e_formatar(df_raw)
            st.session_state["df_dia"] = df
            st.success(f"✅ {len(df)} equipes encontradas com prefixo GOOC/GOOH/GOOL/GOOK")
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")
            return

    if "df_dia" not in st.session_state:
        st.info("Aguardando upload do arquivo.")
        return

    df = st.session_state["df_dia"]

    total = len(df)
    abertos = (df["Turno Aberto"] == "✅ Sim").sum()
    fechados = total - abertos
    com_intervalo = (df["Intervalo"] != "").sum()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total de Equipes", total)
    c2.metric("✅ Turno Aberto", abertos)
    c3.metric("❌ Sem Turno", fechados)
    c4.metric("🔁 Com Intervalo", com_intervalo)

    st.markdown("---")
    st.subheader("📊 Tabela de Turnos")

    def colorir(row):
        if row["Turno Aberto"] == "✅ Sim":
            return ["background-color: #d4edda"] * len(row)
        return ["background-color: #f8d7da"] * len(row)

    st.dataframe(
        df.style.apply(colorir, axis=1),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("---")
    col_gs, col_xlsx, col_csv = st.columns([2, 1, 1])

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    with col_gs:
        if client and sid:
            if st.button("☁️ Salvar no Google Sheets", type="primary", use_container_width=True):
                with st.spinner("Salvando..."):
                    try:
                        # Primeiro verifica o histórico baseando-se no estado atual da planilha, depois salva os novos dados
                        ok2 = atualizar_historico(client, sid, df)
                        ok1 = salvar_turnos_gs(client, sid, df)
                        if ok1 and ok2:
                            st.success("✅ Dados processados e salvos sem duplicidades no histórico!")
                        else:
                            st.error("Erro ao salvar.")
                    except Exception as e:
                        st.error(f"Erro: {e}")
        else:
            st.warning("⚙️ Configure o Google Sheets na aba **Configurações**.")

    with col_xlsx:
        xlsx_bytes = gerar_excel(df)
        st.download_button(
            "📥 Baixar Excel",
            xlsx_bytes,
            f"turnos_{date.today()}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_csv:
        csv = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("📄 Baixar CSV", csv, f"turnos_{date.today()}.csv", "text/csv", use_container_width=True)


def pagina_edicao():
    st.title("✏️ Gerenciamento de Intervalos e Edição")

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    
    if not (client and sid):
        st.warning("Configure o Google Sheets na aba **Configurações**.")
        return

    ws = get_or_create_sheet(client, sid, "Turnos")
    if ws is None:
        return

    df = load_sheet_df(ws)
    if df.empty:
        st.info("Nenhum dado encontrado na planilha.")
        return

    # Divisão em duas abas operacionais
    tab_fechar, tab_atribuir = st.tabs(["🕐 Finalizar Intervalos Pendentes", "➕ Atribuir Início de Intervalo"])

    # ─── ABA 1: FINALIZAR INTERVALOS ABERTOS PENDENTES ───
    with tab_fechar:
        # Regra: Tem que possuir Intervalo registrado, mas Fim de Intervalo precisa estar vazio
        mask_pendentes = (df["Intervalo"].astype(str).str.strip() != "") & \
                         (df["Fim de Intervalo"].astype(str).str.strip() == "")
                         
        df_pendentes = df[mask_pendentes].copy()

        if df_pendentes.empty:
            st.success("🎉 Excelente! Nenhuma equipe possui intervalo pendente de fechamento neste momento.")
        else:
            st.warning(f"⚠️ Há {len(df_pendentes)} equipe(s) com o intervalo em andamento (Sem horário de fim).")
            
            idx_sel = st.selectbox(
                "Selecione a equipe que terminou o intervalo para atualizar:",
                df_pendentes.index,
                format_func=lambda i: f"Data: {df_pendentes.at[i,'Data']} | Equipe: {df_pendentes.at[i,'Prefixo']} (Início do Intervalo às: {df_pendentes.at[i,'Intervalo']})",
                key="sb_pendentes"
            )
            
            row = df.loc[idx_sel]
            
            with st.form("form_fechar_intervalo"):
                st.markdown(f"### Atualizando: **{row['Prefixo']}**")
                c1, c2 = st.columns(2)
                fim_intervalo = c1.text_input("🕐 Definir Fim de Intervalo (HH:MM:SS)", value=datetime.now().strftime("%H:%M:%S"))
                motivo = c2.text_input("📌 Motivo", value=str(row.get("Motivo", "")))
                ocorrencia = st.text_area("📋 Ocorrência", value=str(row.get("Ocorrência", "")), height=80)

                submitted_fim = st.form_submit_button("💾 Confirmar Fim de Intervalo", type="primary")

            if submitted_fim:
                df.at[idx_sel, "Fim de Intervalo"] = fim_intervalo
                df.at[idx_sel, "Motivo"] = motivo
                df.at[idx_sel, "Ocorrência"] = ocorrencia
                try:
                    ws.clear()
                    dados_edicao = [HEADERS_TURNOS]
                    for _, r in df.iterrows():
                        dados_edicao.append([str(r.get(h, "")) for h in HEADERS_TURNOS])
                    ws.append_rows(dados_edicao)
                    st.success("✅ Fim de intervalo registrado com sucesso!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao salvar: {e}")

    # ─── ABA 2: ATRIBUIR NOVO INTERVALO MANUALMENTE ───
    with tab_atribuir:
        st.markdown("Utilize esta seção caso uma equipe tenha saído de intervalo mas o horário não veio registrado automaticamente.")
        
        # Filtra equipes cadastradas que ainda não possuem NENHUM intervalo registrado hoje
        mask_sem_intervalo = (df["Intervalo"].astype(str).str.strip() == "")
        df_sem_intervalo = df[mask_sem_intervalo].copy()
        
        if df_sem_intervalo.empty:
            st.info("Todas as equipes listadas já possuem intervalos atribuídos.")
        else:
            idx_atribuir = st.selectbox(
                "Selecione a equipe para abrir o intervalo:",
                df_sem_intervalo.index,
                format_func=lambda i: f"Data: {df_sem_intervalo.at[i,'Data']} | Equipe: {df_sem_intervalo.at[i,'Prefixo']}",
                key="sb_atribuir"
            )
            
            row_atr = df.loc[idx_atribuir]
            
            with st.form("form_atribuir_intervalo"):
                st.markdown(f"### Atribuir Início para: **{row_atr['Prefixo']}**")
                c1, c2 = st.columns(2)
                inicio_intervalo = c1.text_input("⏳ Definir Horário de Início (HH:MM:SS)", value=datetime.now().strftime("%H:%M:%S"))
                motivo_atr = c2.text_input("📌 Motivo / Justificativa", value="")
                
                submitted_inicio = st.form_submit_button("➕ Abrir Intervalo da Equipe", type="primary")
                
            if submitted_inicio:
                df.at[idx_atribuir, "Intervalo"] = inicio_intervalo
                df.at[idx_atribuir, "Motivo"] = motivo_atr
                
                # Atualiza também a estatística do histórico para registrar que houve um intervalo associado
                ws_hist = get_or_create_sheet(client, sid, "Histórico")
                if ws_hist:
                    df_hist = load_sheet_df(ws_hist)
                    hoje_obj = date.today()
                    mes_m = hoje_obj.strftime("%Y-%m")
                    sem_s = hoje_obj.strftime("%Y-W%W")
                    
                    mask_h = (df_hist.get("Mês", pd.Series()) == mes_m) & \
                             (df_hist.get("Prefixo", pd.Series()) == row_atr['Prefixo']) & \
                             (df_hist.get("Semana", pd.Series()) == sem_s)
                             
                    if not df_hist.empty and mask_h.any():
                        idx_h = df_hist[mask_h].index[0]
                        df_hist.at[idx_h, "Intervalos na Semana"] = int(df_hist.at[idx_h, "Intervalos na Semana"] or 0) + 1
                        df_hist.at[idx_h, "Intervalos no Mês"] = int(df_hist.at[idx_h, "Intervalos no Mês"] or 0) + 1
                        
                        ws_hist.clear()
                        dados_h = [HEADERS_HISTORICO]
                        for _, r in df_hist.iterrows():
                            dados_h.append([str(r.get(h, "")) for h in HEADERS_HISTORICO])
                        ws_hist.append_rows(dados_h)

                try:
                    ws.clear()
                    dados_edicao = [HEADERS_TURNOS]
                    for _, r in df.iterrows():
                        dados_edicao.append([str(r.get(h, "")) for h in HEADERS_TURNOS])
                    ws.append_rows(dados_edicao)
                    st.success(f"✅ Intervalo da equipe {row_atr['Prefixo']} iniciado às {inicio_intervalo}!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao salvar: {e}")

    st.markdown("---")
    st.subheader("📊 Visualização de Todos os Dados da Planilha")
    
    # Filtros de visualização geral
    col1, col2 = st.columns(2)
    datas = sorted(df["Data"].unique(), reverse=True) if "Data" in df.columns else []
    data_sel = col1.selectbox("Filtrar Tabela por Data", ["Todas"] + datas)
    prefixos = sorted(df["Prefixo"].unique()) if "Prefixo" in df.columns else []
    prefixo_sel = col2.selectbox("Filtrar Tabela por Prefixo", ["Todos"] + prefixos)

    df_filtrado = df.copy()
    if data_sel != "Todas":
        df_filtrado = df_filtrado[df_filtrado["Data"] == data_sel]
    if prefixo_sel != "Todos":
        df_filtrado = df_filtrado[df_filtrado["Prefixo"] == prefixo_sel]
        
    st.dataframe(df_filtrado, use_container_width=True, hide_index=True)


def pagina_historico():
    st.title("📈 Histórico de Turnos e Intervalos")

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    
    if not (client and sid):
        st.warning("Configure o Google Sheets na aba **Configurações**.")
        return

    ws = get_or_create_sheet(client, sid, "Histórico")
    if ws is None:
        return

    df = load_sheet_df(ws)
    if df.empty:
        st.info("Nenhum histórico encontrado.")
        return

    meses = sorted(df["Mês"].unique(), reverse=True) if "Mês" in df.columns else []
    mes_sel = st.selectbox("📅 Filtrar por Mês", ["Todos"] + meses)
    if mes_sel != "Todos":
        df = df[df["Mês"] == mes_sel]

    st.subheader("📊 Resumo por Equipe")

    for col in ["Turnos no Mês", "Intervalos na Semana", "Intervalos no Mês"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    st.dataframe(df, use_container_width=True, hide_index=True)

    if not df.empty and "Prefixo" in df.columns and "Turnos no Mês" in df.columns:
        st.markdown("---")
        st.subheader("🏆 Ranking de Turnos no Mês")
        ranking = df.groupby("Prefixo")["Turnos no Mês"].sum().sort_values(ascending=False)
        st.bar_chart(ranking)


def pagina_config():
    st.title("⚙️ Configurações do Google Sheets")

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    if client and sid:
        try:
            ss = client.open_by_key(sid)
            st.success(f"✅ Google Sheets conectado: **{ss.title}**")
        except Exception as e:
            st.error(f"Credenciais encontradas, mas erro ao conectar: {e}")
    else:
        st.warning("Google Sheets não configurado.")

    st.markdown("---")

    with st.expander("📁 Opção 1 — secrets.toml (recomendado para deploy)", expanded=True):
        st.markdown("""
**Passo a passo:**
**1.** Crie a pasta `.streamlit` na raiz do projeto e o arquivo `secrets.toml` dentro dela.
**2.** Abra o `secrets.toml` e cole com as três aspas na private_key conforme o padrão TOML.
        """)

    with st.expander("🔑 Opção 2 — Configuração manual (apenas nesta sessão)"):
        creds_json = st.text_area(
            "Conteúdo do JSON de credenciais:",
            value=json.dumps(st.session_state.get("gsheet_creds", {}), indent=2) if st.session_state.get("gsheet_creds") else "",
            height=220,
        )

        spreadsheet_id_input = st.text_input(
            "ID da planilha:",
            value=st.session_state.get("spreadsheet_id", ""),
        )

        col1, col2 = st.columns(2)
        if col1.button("💾 Salvar e testar", type="primary"):
            if creds_json.strip():
                try:
                    creds_dict = json.loads(creds_json)
                    st.session_state["gsheet_creds"] = creds_dict
                    st.session_state["spreadsheet_id"] = spreadsheet_id_input.strip()
                    c = get_gsheet_client()
                    ss = c.open_by_key(spreadsheet_id_input.strip())
                    st.success(f"✅ Conectado: **{ss.title}**")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao conectar: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/maintenance.png", width=60)
        st.title("Controle de Turnos")
        st.markdown("---")
        pagina = st.radio(
            "Navegação",
            ["📋 Turnos do Dia", "✏️ Editar Registros", "📈 Histórico", "⚙️ Configurações"],
        )
        st.markdown("---")
        st.caption(f"Data: {date.today().strftime('%%d/%%m/%%Y')}")
        
        if get_gsheet_client() and get_spreadsheet_id():
            st.success("Google Sheets ✅")
        else:
            st.warning("Google Sheets ⚠️")

    if pagina == "📋 Turnos do Dia":
        pagina_upload()
    elif pagina == "✏️ Editar Registros":
        pagina_edicao()
    elif pagina == "📈 Histórico":
        pagina_historico()
    elif pagina == "⚙️ Configurações":
        pagina_config()


if __name__ == "__main__":
    main()
