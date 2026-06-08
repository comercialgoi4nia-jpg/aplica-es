import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, date
import json
import io

st.set_page_config(page_title="Controle de Turnos", page_icon="🔧", layout="wide")

PREFIXOS_ALVO = ("GOOC", "GOOH", "GOOL", "GOOK")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ─── Excel download ──────────────────────────────────────────────────────────

def gerar_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos"

    verde_fill = PatternFill("solid", start_color="C6EFCE")
    vermelho_fill = PatternFill("solid", start_color="FFC7CE")
    header_fill = PatternFill("solid", start_color="2E4057")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell_font = Font(name="Arial", size=10)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    ws.row_dimensions[1].height = 22

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        aberto = str(row.get("Turno Aberto", "")) == "✅ Sim"
        fill = verde_fill if aberto else vermelho_fill
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(row[h]))
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda

    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), *(len(str(row[h])) for _, row in df.iterrows()))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Google Sheets (CORRIGIDO) ────────────────────────────────────────────────

def get_gsheet_client():
    creds_dict = st.session_state.get("gsheet_creds")
    if not creds_dict:
        try:
            # CORREÇÃO: Usa .to_dict() para ler os segredos estruturados do Streamlit
            creds_dict = st.secrets["gcp_service_account"].to_dict()
        except Exception:
            return None
            
    try:
        # CORREÇÃO: Garante que as quebras de linha da chave privada sejam interpretadas textualmente
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
            
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        return gspread.authorize(creds)
    except Exception as e:
        st.sidebar.error(f"Erro nas credenciais: {e}")
        return None


def get_spreadsheet_id():
    sid = st.session_state.get("spreadsheet_id", "")
    if not sid:
        try:
            sid = st.secrets["spreadsheet_id"]
        except Exception:
            pass
    return str(sid).strip()


def get_or_create_sheet(client, spreadsheet_id, sheet_name):
    try:
        ss = client.open_by_key(spreadsheet_id)
    except Exception as e:
        st.error(f"Erro ao abrir planilha: {e}")
        return None
    try:
        return ss.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=sheet_name, rows=1000, cols=20)
        return ws


def ensure_headers(ws, headers):
    existing = ws.row_values(1)
    if existing != headers:
        ws.clear()
        ws.append_row(headers)


def load_sheet_df(ws):
    data = ws.get_all_records(default_blank="")
    return pd.DataFrame(data) if data else pd.DataFrame(columns=ws.row_values(1))


# ─── XML / Excel reader ───────────────────────────────────────────────────────

def ler_xml(conteudo: bytes) -> pd.DataFrame:
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError as e:
        raise ValueError(f"XML inválido: {e}")

    NS = "urn:schemas-microsoft-com:office:spreadsheet"

    def tag(nome):
        return f"{{{NS}}}{nome}"

    worksheet = root.find(f".//{tag('Worksheet')}")
    if worksheet is not None:
        table = worksheet.find(tag("Table"))
        if table is None:
            for ws in root.iter(tag("Worksheet")):
                table = ws.find(tag("Table"))
                if table is not None:
                    break
        if table is not None:
            linhas = list(table.findall(tag("Row")))
            if not linhas:
                raise ValueError("SpreadsheetML: nenhuma linha encontrada.")
            ATTR_INDEX = f"{{{NS}}}Index"

            def celulas(row_el):
                vals = []
                idx = 0
                for cell in row_el.findall(tag("Cell")):
                    ss_idx = cell.get(ATTR_INDEX)
                    if ss_idx:
                        alvo = int(ss_idx) - 1
                        while idx < alvo:
                            vals.append("")
                            idx += 1
                    data_el = cell.find(tag("Data"))
                    vals.append((data_el.text or "").strip() if data_el is not None else "")
                    idx += 1
                return vals

            cabecalho = celulas(linhas[0])
            cab_norm = []
            contagem: dict = {}
            for c in cabecalho:
                c = c.strip() if c.strip() else f"_col{len(cab_norm)}"
                contagem[c] = contagem.get(c, 0) + 1
                cab_norm.append(c if contagem[c] == 1 else f"{c}_{contagem[c]}")

            registros = []
            for row_el in linhas[1:]:
                vals = celulas(row_el)
                while len(vals) < len(cab_norm):
                    vals.append("")
                registros.append(dict(zip(cab_norm, vals[: len(cab_norm)])))

            if not registros:
                raise ValueError("SpreadsheetML: tabela sem linhas de dados.")
            df = pd.DataFrame(registros)
            df.columns = [str(c).strip() for c in df.columns]
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)].reset_index(drop=True)
            return df

    filhos = list(root)
    if not filhos:
        raise ValueError("XML sem elementos filhos na raiz.")
    candidatos = filhos
    for _ in range(3):
        netos = []
        for f in candidatos:
            netos.extend(list(f))
        if len(netos) > len(candidatos):
            candidatos = netos
        else:
            break
    amostra = candidatos[0]
    registros = []
    if list(amostra):
        def limpar_tag(t):
            return t.split("}")[-1] if "}" in t else t
        for filho in candidatos:
            registro = {limpar_tag(sub.tag): (sub.text or "").strip() for sub in filho}
            if registro:
                registros.append(registro)
    elif amostra.attrib:
        for filho in candidatos:
            if filho.attrib:
                registros.append(dict(filho.attrib))
    if not registros:
        raise ValueError("Não foi possível extrair registros do XML.")
    df = pd.DataFrame(registros)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def ler_arquivo(arquivo) -> pd.DataFrame:
    nome = arquivo.name.lower()
    conteudo = arquivo.read()
    arquivo.seek(0)
    if nome.endswith(".xml"):
        return ler_xml(conteudo)
    elif nome.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(conteudo))
    else:
        raise ValueError(f"Formato não suportado: {nome}")


def filtrar_e_formatar(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["Prefixo"].astype(str).str.startswith(PREFIXOS_ALVO)
    df = df[mask].copy()

    def fmt_hora(val):
        if pd.isna(val) or str(val).strip() in ("", "NaT"):
            return ""
        try:
            if isinstance(val, str):
                dt = pd.to_datetime(val, errors="coerce")
            else:
                dt = pd.Timestamp(val)
            if pd.isna(dt):
                return str(val)
            return dt.strftime("%H:%M:%S")
        except Exception:
            return str(val)

    result = pd.DataFrame()
    result["Prefixo"] = df["Prefixo"].astype(str)
    result["Início de Turno"] = df["Início de Turno"].apply(fmt_hora)
    result["Previsão de Saída"] = df["Previsão Saída"].apply(fmt_hora)
    result["Intervalo"] = df["Intervalo"].apply(fmt_hora)
    result["Turno Aberto"] = result["Início de Turno"].apply(lambda x: "✅ Sim" if x else "❌ Não")
    result["Data"] = date.today().strftime("%Y-%m-%d")
    return result.reset_index(drop=True)


# ─── Google Sheets: Salvar e Otimizar (CORRIGIDO SEM ERRO 429) ─────────────────

HEADERS_TURNOS = ["Data", "Prefixo", "Início de Turno", "Previsão de Saída", "Intervalo",
                  "Fim de Intervalo", "Motivo", "Ocorrência", "Turno Aberto"]
HEADERS_HISTORICO = ["Mês", "Semana", "Prefixo", "Turnos no Mês", "Intervalos na Semana",
                     "Intervalos no Mês"]


def salvar_turnos_gs(client, spreadsheet_id, df_dia):
    ws = get_or_create_sheet(client, spreadsheet_id, "Turnos")
    if ws is None:
        return False
    ensure_headers(ws, HEADERS_TURNOS)
    df_atual = load_sheet_df(ws)
    hoje = date.today().strftime("%Y-%m-%d")
    
    if not df_atual.empty and "Data" in df_atual.columns:
        df_atual = df_atual[df_atual["Data"].astype(str) != hoje]
        
    novos = []
    for _, row in df_dia.iterrows():
        novos.append({
            "Data": row["Data"],
            "Prefixo": row["Prefixo"],
            "Início de Turno": row["Início de Turno"],
            "Previsão de Saída": row["Previsão de Saída"],
            "Intervalo": row["Intervalo"],
            "Fim de Intervalo": "",
            "Motivo": "",
            "Ocorrência": "",
            "Turno Aberto": row["Turno Aberto"],
        })
    
    df_final = pd.concat([df_atual, pd.DataFrame(novos)], ignore_index=True)
    ws.clear()
    
    # CORREÇÃO: Prepara toda a tabela em memória primeiro
    dados_para_enviar = [HEADERS_TURNOS]
    for _, r in df_final.iterrows():
        dados_para_enviar.append([str(r.get(h, "")) for h in HEADERS_TURNOS])
    
    # Envia o bloco inteiro em 1 só requisição (Otimizado)
    ws.append_rows(dados_para_enviar) 
    return True


def atualizar_historico(client, spreadsheet_id, df_dia):
    ws = get_or_create_sheet(client, spreadsheet_id, "Histórico")
    if ws is None:
        return False
    ensure_headers(ws, HEADERS_HISTORICO)
    df_hist = load_sheet_df(ws)
    hoje = date.today()
    mes_atual = hoje.strftime("%Y-%m")
    semana_atual = hoje.strftime("%Y-W%W")

    for _, row in df_dia.iterrows():
        prefixo = row["Prefixo"]
        abriu = row["Turno Aberto"] == "✅ Sim"
        tem_intervalo = bool(row["Intervalo"])

        mask_mes = (df_hist.get("Mês", pd.Series()) == mes_atual) & \
                   (df_hist.get("Prefixo", pd.Series()) == prefixo) & \
                   (df_hist.get("Semana", pd.Series()) == semana_atual)

        if df_hist.empty or not mask_mes.any():
            df_hist = pd.concat([df_hist, pd.DataFrame([{
                "Mês": mes_atual,
                "Semana": semana_atual,
                "Prefixo": prefixo,
                "Turnos no Mês": 1 if abriu else 0,
                "Intervalos na Semana": 1 if tem_intervalo else 0,
                "Intervalos no Mês": 1 if tem_intervalo else 0,
            }])], ignore_index=True)
        else:
            idx = df_hist[mask_mes].index[0]
            if abriu:
                df_hist.at[idx, "Turnos no Mês"] = int(df_hist.at[idx, "Turnos no Mês"] or 0) + 1
            if tem_intervalo:
                df_hist.at[idx, "Intervalos na Semana"] = int(df_hist.at[idx, "Intervalos na Semana"] or 0) + 1
                df_hist.at[idx, "Intervalos no Mês"] = int(df_hist.at[idx, "Intervalos no Mês"] or 0) + 1

    ws.clear()
    
    # CORREÇÃO: Prepara o bloco todo do Histórico
    dados_historico = [HEADERS_HISTORICO]
    for _, r in df_hist.iterrows():
        dados_historico.append([str(r.get(h, "")) for h in HEADERS_HISTORICO])
        
    # Envia tudo de uma vez só
    ws.append_rows(dados_historico)
    return True


# ─── Páginas ──────────────────────────────────────────────────────────────────

def pagina_upload():
    st.title("📋 Controle de Turnos do Dia")
    st.markdown("Faça o upload da base (XML ou Excel) para visualizar e registrar os turnos.")

    arquivo = st.file_uploader("📂 Selecione o arquivo da base", type=["xml", "xlsx", "xls"])

    if arquivo:
        try:
            df_raw = ler_arquivo(arquivo)
            df = filtrar_e_formatar(df_raw)
            st.session_state["df_dia"] = df
            st.success(f"✅ {len(df)} equipes encontradas com prefixo GOOC/GOOH/GOOL/GOOK")
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")
            return

    if "df_dia" not in st.session_state:
        st.info("Aguardando upload do arquivo.")
        return

    df = st.session_state["df_dia"]

    total = len(df)
    abertos = (df["Turno Aberto"] == "✅ Sim").sum()
    fechados = total - abertos
    com_intervalo = (df["Intervalo"] != "").sum()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total de Equipes", total)
    c2.metric("✅ Turno Aberto", abertos)
    c3.metric("❌ Sem Turno", fechados)
    c4.metric("🔁 Com Intervalo", com_intervalo)

    st.markdown("---")
    st.subheader("📊 Tabela de Turnos")

    def colorir(row):
        if row["Turno Aberto"] == "✅ Sim":
            return ["background-color: #d4edda"] * len(row)
        return ["background-color: #f8d7da"] * len(row)

    st.dataframe(
        df.style.apply(colorir, axis=1),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("---")
    col_gs, col_xlsx, col_csv = st.columns([2, 1, 1])

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    with col_gs:
        if client and sid:
            if st.button("☁️ Salvar no Google Sheets", type="primary", use_container_width=True):
                with st.spinner("Salvando..."):
                    try:
                        ok1 = salvar_turnos_gs(client, sid, df)
                        ok2 = atualizar_historico(client, sid, df)
                        if ok1 and ok2:
                            st.success("✅ Dados salvos no Google Sheets!")
                        else:
                            st.error("Erro ao salvar.")
                    except Exception as e:
                        st.error(f"Erro: {e}")
        else:
            st.warning("⚙️ Configure o Google Sheets na aba **Configurações**.")

    with col_xlsx:
        xlsx_bytes = gerar_excel(df)
        st.download_button(
            "📥 Baixar Excel",
            xlsx_bytes,
            f"turnos_{date.today()}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_csv:
        csv = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("📄 Baixar CSV", csv, f"turnos_{date.today()}.csv", "text/csv", use_container_width=True)


def pagina_edicao():
    st.title("✏️ Editar Registros de Turno")

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    
    if not (client and sid):
        st.warning("Configure o Google Sheets na aba **Configurações**.")
        return

    ws = get_or_create_sheet(client, sid, "Turnos")
    if ws is None:
        return

    df = load_sheet_df(ws)
    if df.empty:
        st.info("Nenhum dado encontrado na planilha.")
        return

    col1, col2 = st.columns(2)
    datas = sorted(df["Data"].unique(), reverse=True) if "Data" in df.columns else []
    data_sel = col1.selectbox("Filtrar por Data", ["Todas"] + datas)
    prefixos = sorted(df["Prefixo"].unique()) if "Prefixo" in df.columns else []
    prefixo_sel = col2.selectbox("Filtrar por Prefixo", ["Todos"] + prefixos)

    df_filtrado = df.copy()
    if data_sel != "Todas":
        df_filtrado = df_filtrado[df_filtrado["Data"] == data_sel]
    if prefixo_sel != "Todos":
        df_filtrado = df_filtrado[df_filtrado["Prefixo"] == prefixo_sel]

    st.subheader(f"📝 {len(df_filtrado)} registro(s) encontrado(s)")

    if df_filtrado.empty:
        st.info("Nenhum registro para os filtros selecionados.")
        return

    idx_sel = st.selectbox(
        "Selecione o registro para editar",
        df_filtrado.index,
        format_func=lambda i: f"{df_filtrado.at[i,'Data']} | {df_filtrado.at[i,'Prefixo']}",
    )
    row = df.loc[idx_sel]

    st.markdown("---")
    with st.form("form_edicao"):
        c1, c2 = st.columns(2)
        fim_intervalo = c1.text_input("🕐 Fim de Intervalo (HH:MM:SS)", value=str(row.get("Fim de Intervalo", "")))
        motivo = c2.text_input("📌 Motivo", value=str(row.get("Motivo", "")))
        ocorrencia = st.text_area("📋 Ocorrência", value=str(row.get("Ocorrência", "")), height=100)

        submitted = st.form_submit_button("💾 Salvar Alterações", type="primary")

    if submitted:
        df.at[idx_sel, "Fim de Intervalo"] = fim_intervalo
        df.at[idx_sel, "Motivo"] = motivo
        df.at[idx_sel, "Ocorrência"] = ocorrencia
        try:
            ws.clear()
            
            # CORREÇÃO: Aplica a otimização de bloco no salvamento da edição também
            dados_edicao = [HEADERS_TURNOS]
            for _, r in df.iterrows():
                dados_edicao.append([str(r.get(h, "")) for h in HEADERS_TURNOS])
                
            ws.append_rows(dados_edicao)
            st.success("✅ Registro atualizado com sucesso!")
            st.rerun()
        except Exception as e:
            st.error(f"Erro ao salvar: {e}")

    st.markdown("---")
    st.subheader("📊 Visualização completa")
    st.dataframe(df_filtrado, use_container_width=True, hide_index=True)


def pagina_historico():
    st.title("📈 Histórico de Turnos e Intervalos")

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    
    if not (client and sid):
        st.warning("Configure o Google Sheets na aba **Configurações**.")
        return

    ws = get_or_create_sheet(client, sid, "Histórico")
    if ws is None:
        return

    df = load_sheet_df(ws)
    if df.empty:
        st.info("Nenhum histórico encontrado.")
        return

    meses = sorted(df["Mês"].unique(), reverse=True) if "Mês" in df.columns else []
    mes_sel = st.selectbox("📅 Filtrar por Mês", ["Todos"] + meses)
    if mes_sel != "Todos":
        df = df[df["Mês"] == mes_sel]

    st.subheader("📊 Resumo por Equipe")

    for col in ["Turnos no Mês", "Intervalos na Semana", "Intervalos no Mês"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    st.dataframe(df, use_container_width=True, hide_index=True)

    if not df.empty and "Prefixo" in df.columns and "Turnos no Mês" in df.columns:
        st.markdown("---")
        st.subheader("🏆 Ranking de Turnos no Mês")
        ranking = df.groupby("Prefixo")["Turnos no Mês"].sum().sort_values(ascending=False)
        st.bar_chart(ranking)


def pagina_config():
    st.title("⚙️ Configurações do Google Sheets")

    client = get_gsheet_client()
    sid = get_spreadsheet_id()
    if client and sid:
        try:
            ss = client.open_by_key(sid)
            st.success(f"✅ Google Sheets conectado: **{ss.title}**")
        except Exception as e:
            st.error(f"Credenciais encontradas, mas erro ao conectar: {e}")
    else:
        st.warning("Google Sheets não configurado.")

    st.markdown("---")

    with st.expander("📁 Opção 1 — secrets.toml (recomendado para deploy)", expanded=True):
        st.markdown("""
**Passo a passo:**

**1.** Crie a pasta `.streamlit` na raiz do projeto e o arquivo `secrets.toml` dentro dela.
**2.** Abra o `secrets.toml` e cole com as três aspas na private_key.
        """)

    with st.expander("🔑 Opção 2 — Configuração manual (apenas nesta sessão)"):
        creds_json = st.text_area(
            "Conteúdo do JSON de credenciais:",
            value=json.dumps(st.session_state.get("gsheet_creds", {}), indent=2) if st.session_state.get("gsheet_creds") else "",
            height=220,
        )

        spreadsheet_id_input = st.text_input(
            "ID da planilha:",
            value=st.session_state.get("spreadsheet_id", ""),
        )

        col1, col2 = st.columns(2)
        if col1.button("💾 Salvar e testar", type="primary"):
            if creds_json.strip():
                try:
                    creds_dict = json.loads(creds_json)
                    st.session_state["gsheet_creds"] = creds_dict
                    st.session_state["spreadsheet_id"] = spreadsheet_id_input.strip()
                    c = get_gsheet_client()
                    ss = c.open_by_key(spreadsheet_id_input.strip())
                    st.success(f"✅ Conectado: **{ss.title}**")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao conectar: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/maintenance.png", width=60)
        st.title("Controle de Turnos")
        st.markdown("---")
        pagina = st.radio(
            "Navegação",
            ["📋 Turnos do Dia", "✏️ Editar Registros", "📈 Histórico", "⚙️ Configurações"],
        )
        st.markdown("---")
        st.caption(f"Data: {date.today().strftime('%d/%m/%Y')}")
        
        # Correção visual do indicador na barra lateral
        if get_gsheet_client() and get_spreadsheet_id():
            st.success("Google Sheets ✅")
        else:
            st.warning("Google Sheets ⚠️")

    if pagina == "📋 Turnos do Dia":
        pagina_upload()
    elif pagina == "✏️ Editar Registros":
        pagina_edicao()
    elif pagina == "📈 Histórico":
        pagina_historico()
    elif pagina == "⚙️ Configurações":
        pagina_config()


if __name__ == "__main__":
    main()
