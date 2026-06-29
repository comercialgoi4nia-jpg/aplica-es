import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Conferência MISCELÂNEA", page_icon="📋", layout="wide")

st.markdown("""
    <style>
    .main { background-color: #f5f7fa; }
    .stButton>button { background-color: #1a73e8; color: white; border-radius: 8px; font-weight: bold; }
    .metric-card { background: white; padding: 16px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    </style>
""", unsafe_allow_html=True)

st.title("📋 Conferência de MISCELÂNEA")
st.markdown("Compare a coluna **MISCELÂNEA** entre a base de abertas e a base de executadas.")

st.divider()

col1, col2 = st.columns(2)

with col1:
    st.subheader("📂 Base Abertas")
    st.caption("Planilha com os chamados abertos (referência)")
    file_abertas = st.file_uploader("Selecione o arquivo", type=["xlsx", "xls", "csv"], key="abertas")

with col2:
    st.subheader("✅ Base Executadas")
    st.caption("Planilha com os chamados já executados")
    file_executadas = st.file_uploader("Selecione o arquivo", type=["xlsx", "xls", "csv"], key="executadas")

COLUNAS_ABERTAS = [
    "Bloco", "N° Ticket", "Município", "Qtd Câmeras",
    "Carga Inst. (W)", "Consumo (kWh/mês)", "Latitude", "Ponto",
    "Longitude", "PONTO DE SERVIÇO", "MISCELANEA", "Endereço"
]

def load_file(f):
    if f.name.endswith(".csv"):
        return pd.read_csv(f, dtype=str)
    return pd.read_excel(f, dtype=str)

def normalize(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    return str(val).strip().upper()

def normalizar_nome_col(nome):
    """Remove acentos, espaços extras e converte para maiúsculo."""
    import unicodedata
    import re
    s = str(nome).strip()
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    return s

def encontrar_col_miscelanea(df):
    """Encontra coluna cujo nome normalizado contenha MISCELANEA."""
    for col in df.columns:
        norm = normalizar_nome_col(col)
        if "MISCELANEA" in norm:
            return col
    return None

def gerar_excel(df_resultado, df_abrir_misc=None, col_end_ex=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Conferência"

    # Colors
    verde_fill   = PatternFill("solid", fgColor="C6EFCE")  # executada
    vermelho_fill = PatternFill("solid", fgColor="FFC7CE")  # não executada
    amarelo_fill  = PatternFill("solid", fgColor="FFEB9C")  # abrir miscelanea
    header_fill  = PatternFill("solid", fgColor="1A73E8")

    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = list(df_resultado.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    status_col = headers.index("STATUS") + 1

    for row_idx, row in enumerate(df_resultado.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

        status = ws.cell(row=row_idx, column=status_col).value
        if status == "EXECUTADA":
            fill = verde_fill
            ws.cell(row=row_idx, column=status_col).font = Font(bold=True, color="375623")
        elif status == "NÃO EXECUTADA":
            fill = vermelho_fill
            ws.cell(row=row_idx, column=status_col).font = Font(bold=True, color="9C0006")
        else:
            fill = PatternFill()

        # Mark MISCELANEA col if "ABRIR MISCELANEA"
        misc_col_idx = headers.index("MISCELANEA_EXECUTADAS") + 1 if "MISCELANEA_EXECUTADAS" in headers else None

        for col_idx2 in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx2).fill = fill

        if misc_col_idx:
            misc_val = ws.cell(row=row_idx, column=misc_col_idx).value
            if misc_val == "ABRIR MISCELANEA":
                ws.cell(row=row_idx, column=misc_col_idx).fill = amarelo_fill
                ws.cell(row=row_idx, column=misc_col_idx).font = Font(bold=True, color="7B5E00")

    # Auto-width
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Aba ABRIR MISCELANEA: linhas da executada sem miscelânea preenchida
    if df_abrir_misc is not None and len(df_abrir_misc) > 0:
        ws_ab = wb.create_sheet("ABRIR MISCELANEA")
        # Escolhe colunas relevantes: miscelanea + endereço + demais disponíveis
        cols_show = list(df_abrir_misc.columns)
        # Garante endereço aparece em destaque
        if col_end_ex and col_end_ex in cols_show:
            cols_show = [col_end_ex] + [c for c in cols_show if c != col_end_ex and c != "MISCELANEA_STATUS"]
        else:
            cols_show = [c for c in cols_show if c != "MISCELANEA_STATUS"]
        df_show_ab = df_abrir_misc[cols_show]

        # Cabeçalho
        for ci, h in enumerate(df_show_ab.columns, 1):
            cell = ws_ab.cell(row=1, column=ci, value=h)
            cell.fill = PatternFill("solid", fgColor="FF9900")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
        # Dados
        for ri, row in enumerate(df_show_ab.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                cell = ws_ab.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
        # Auto-width
        for ci, col_cells in enumerate(ws_ab.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws_ab.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 45)
        ws_ab.row_dimensions[1].height = 30
        ws_ab.freeze_panes = "A2"

    # Legend sheet
    ws2 = wb.create_sheet("Legenda")
    ws2["A1"] = "Legenda de Cores"
    ws2["A1"].font = Font(bold=True, size=13)
    legenda = [
        ("Verde", "C6EFCE", "Executada — MISCELÂNEA encontrada na base de executadas"),
        ("Vermelho", "FFC7CE", "Não Executada — MISCELÂNEA não encontrada na base de executadas"),
        ("Amarelo", "FFEB9C", "ABRIR MISCELANEA — linha da executada sem MISCELÂNEA preenchida"),
    ]
    for i, (label, color, desc) in enumerate(legenda, 3):
        ws2.cell(row=i, column=1, value=label).fill = PatternFill("solid", fgColor=color)
        ws2.cell(row=i, column=1).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=desc)
        ws2.column_dimensions["B"].width = 60

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if file_abertas and file_executadas:
    try:
        df_ab = load_file(file_abertas)
        df_ex = load_file(file_executadas)

        # Normalize column names
        df_ab.columns = df_ab.columns.str.strip()
        df_ex.columns = df_ex.columns.str.strip()

        if "MISCELANEA" not in df_ab.columns:
            st.error("❌ Coluna **MISCELANEA** não encontrada na base de Abertas. Verifique o arquivo.")
            st.stop()
        col_misc_ex = encontrar_col_miscelanea(df_ex)
        if col_misc_ex is None:
            st.error("❌ Nenhuma coluna com 'Miscelânea' encontrada na base de Executadas. Verifique o arquivo.")
            st.stop()
        if col_misc_ex != "MISCELANEA":
            st.info(f"ℹ️ Coluna identificada na base de Executadas: **\"{col_misc_ex}\"** → tratada como MISCELANEA.")
            df_ex = df_ex.rename(columns={col_misc_ex: "MISCELANEA"})

        # Lista de valores normalizados da coluna miscelanea nas executadas (apenas não-vazios)
        exec_misc_lista = [
            normalize(v) for v in df_ex["MISCELANEA"]
            if normalize(v) != ""
        ]

        # Linhas da executada com miscelanea vazia recebem marcação ABRIR MISCELANEA
        df_ex["MISCELANEA_STATUS"] = df_ex["MISCELANEA"].apply(
            lambda v: "ABRIR MISCELANEA" if normalize(v) == "" else normalize(v)
        )

        def contem_miscelanea(val_ab):
            """Retorna True se val_ab (aberta) está contido em algum valor da executada ou vice-versa."""
            n = normalize(val_ab)
            if n == "":
                return False
            for ex in exec_misc_lista:
                if n in ex or ex in n:
                    return True
            return False

        # Detectar coluna Endereço na base de executadas (case-insensitive)
        col_end_ex = next(
            (c for c in df_ex.columns if "endere" in normalizar_nome_col(c).lower()),
            None
        )

        def buscar_match_executada(val_ab):
            """Retorna (valor_miscelanea, endereco) da executada que fez match, ou (None, None)."""
            n = normalize(val_ab)
            if n == "":
                return None, None
            for _, row in df_ex.iterrows():
                ex_norm = normalize(row["MISCELANEA"])
                if ex_norm == "":
                    continue
                if n in ex_norm or ex_norm in n:
                    end = str(row[col_end_ex]).strip() if col_end_ex and col_end_ex in row.index else ""
                    return row["MISCELANEA"], end
            return None, None

        # Build result from abertas
        cols_abertas = [c for c in COLUNAS_ABERTAS if c in df_ab.columns]
        df_resultado = df_ab[cols_abertas].copy()

        # Status: EXECUTADA se contém match, senão NÃO EXECUTADA
        df_resultado["STATUS"] = df_resultado["MISCELANEA"].apply(
            lambda v: "EXECUTADA" if contem_miscelanea(v) else "NÃO EXECUTADA"
        )

        # Coluna com o valor da executada que fez match + endereço
        matches = df_resultado["MISCELANEA"].apply(
            lambda v: buscar_match_executada(v)
        )
        df_resultado["MISCELANEA_EXECUTADAS"] = matches.apply(
            lambda t: t[0] if t[0] else ("ABRIR MISCELANEA" if True else "NÃO ENCONTRADA")
        )
        # Corrigir: NÃO ENCONTRADA quando a aberta não está vazia mas não teve match
        def status_misc(v, t):
            if t[0]:
                return t[0]
            if normalize(v) == "":
                return "ABRIR MISCELANEA"
            return "NÃO ENCONTRADA"
        df_resultado["MISCELANEA_EXECUTADAS"] = [
            status_misc(v, t)
            for v, t in zip(df_resultado["MISCELANEA"], matches)
        ]
        # Endereço da executada para linhas com ABRIR MISCELANEA
        # Pega endereços das linhas da executada que têm miscelanea vazia
        if col_end_ex:
            enderecos_abrir = df_ex[df_ex["MISCELANEA_STATUS"] == "ABRIR MISCELANEA"][col_end_ex].dropna().tolist()
        else:
            enderecos_abrir = []

        # Coluna ENDEREÇO_ABRIR_MISCELANEA: preenche com endereço do match quando executou,
        # ou com os endereços das linhas sem miscelânea (ABRIR MISCELANEA) para não executadas
        def obter_endereco(v, t, idx_row):
            if t[0]:
                # executada: endereço da linha que fez match
                return t[1]
            if normalize(v) == "":
                return ""
            return ""
        df_resultado["ENDEREÇO_ABRIR_MISCELANEA"] = [
            obter_endereco(v, t, i)
            for i, (v, t) in enumerate(zip(df_resultado["MISCELANEA"], matches))
        ]

        # Adiciona aba extra no Excel com as linhas ABRIR MISCELANEA da base executadas
        df_abrir_misc = df_ex[df_ex["MISCELANEA_STATUS"] == "ABRIR MISCELANEA"].copy()

        # Conta linhas com ABRIR MISCELANEA (vazias na executada)
        abrir_misc_count = len(df_abrir_misc)

        # Reorder columns: put STATUS first
        cols_order = ["STATUS"] + [c for c in df_resultado.columns if c != "STATUS"]
        df_resultado = df_resultado[cols_order]

        st.divider()
        st.subheader("📊 Resumo")

        total = len(df_resultado)
        executadas = (df_resultado["STATUS"] == "EXECUTADA").sum()
        nao_executadas = (df_resultado["STATUS"] == "NÃO EXECUTADA").sum()
        abrir_misc = abrir_misc_count

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total (Abertas)", total)
        c2.metric("✅ Executadas", executadas, delta=f"{executadas/total*100:.1f}%")
        c3.metric("❌ Não Executadas", nao_executadas, delta=f"-{nao_executadas/total*100:.1f}%", delta_color="inverse")
        c4.metric("⚠️ Abrir Miscelânea", abrir_misc)

        st.divider()
        st.subheader("📋 Resultado")

        filtro = st.radio("Filtrar por:", ["Todos", "Executadas", "Não Executadas"], horizontal=True)
        if filtro == "Executadas":
            df_show = df_resultado[df_resultado["STATUS"] == "EXECUTADA"]
        elif filtro == "Não Executadas":
            df_show = df_resultado[df_resultado["STATUS"] == "NÃO EXECUTADA"]
        else:
            df_show = df_resultado

        def highlight_status(row):
            if row["STATUS"] == "EXECUTADA":
                return ["background-color: #C6EFCE"] * len(row)
            elif row["STATUS"] == "NÃO EXECUTADA":
                return ["background-color: #FFC7CE"] * len(row)
            return [""] * len(row)

        st.dataframe(
            df_show.style.apply(highlight_status, axis=1),
            use_container_width=True,
            height=420
        )

        st.divider()
        excel_bytes = gerar_excel(df_resultado)
        st.download_button(
            label="⬇️ Baixar Planilha com Resultado",
            data=excel_bytes,
            file_name="conferencia_miscelanea.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    except Exception as e:
        st.error(f"Erro ao processar os arquivos: {e}")
        st.exception(e)
else:
    st.info("📎 Carregue os dois arquivos acima para iniciar a conferência.")
    with st.expander("ℹ️ Como funciona o aplicativo"):
        st.markdown("""
        1. **Carregue a Base de Abertas** — planilha com os chamados abertos, deve conter a coluna `MISCELANEA` e as demais colunas obrigatórias.
        2. **Carregue a Base de Executadas** — planilha com os chamados já executados, também com a coluna `MISCELANEA`.
        3. O app compara os valores de `MISCELANEA` entre as duas bases.
        4. Linhas da base de executadas com `MISCELANEA` vazia receberão o valor **ABRIR MISCELANEA**.
        5. O resultado é exibido na tela e pode ser baixado como planilha Excel colorida:
           - 🟢 **Verde** = Executada
           - 🔴 **Vermelho** = Não Executada
           - 🟡 **Amarelo** = Abrir Miscelânea (sem valor na executada)
        """)
