import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Relatório de Serviços Prioritários",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main-header {
        background: linear-gradient(135deg, #0d2b6b 0%, #1a52b3 60%, #2176ff 100%);
        border-radius: 14px; padding: 2rem 2.5rem; margin-bottom: 1.5rem;
        color: white; box-shadow: 0 8px 32px rgba(13,43,107,0.25);
    }
    .main-header h1 { font-size: 2rem; font-weight: 700; margin: 0 0 0.3rem 0; letter-spacing: -0.5px; }
    .main-header p { font-size: 0.95rem; opacity: 0.85; margin: 0; }
    .metric-card {
        background: white; border: 1px solid #e2e8f0; border-radius: 10px;
        padding: 1rem 1.25rem; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .metric-card .label { font-size: 0.75rem; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: 0.6px; }
    .metric-card .value { font-size: 1.8rem; font-weight: 700; color: #0d2b6b; line-height: 1.2; }
    .badge { display: inline-block; padding: 0.25rem 0.75rem; border-radius: 20px; font-size: 0.75rem; font-weight: 600; background: #e8f0fe; color: #1a52b3; margin-bottom: 1rem; }
    .info-box { background: #f0f7ff; border-left: 4px solid #2176ff; border-radius: 0 8px 8px 0; padding: 0.85rem 1.1rem; font-size: 0.9rem; color: #1e3a5f; margin-bottom: 1rem; }
    div[data-testid="stSidebar"] { background: #f8fafc; border-right: 1px solid #e2e8f0; }
    div[data-testid="stSidebar"] .sidebar-title { font-size: 0.7rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; color: #94a3b8; padding: 0.5rem 0 0.25rem 0; }
    .stDownloadButton > button {
        width: 100%; background: linear-gradient(135deg, #0d2b6b, #2176ff) !important;
        color: white !important; border: none !important; border-radius: 8px !important;
        font-weight: 600 !important; padding: 0.65rem 1rem !important; font-size: 0.95rem !important;
        transition: opacity 0.2s; box-shadow: 0 4px 12px rgba(33,118,255,0.3) !important;
    }
    .stDownloadButton > button:hover { opacity: 0.9; }
    .stSelectbox label, .stNumberInput label, .stFileUploader label { font-weight: 600 !important; font-size: 0.85rem !important; color: #334155 !important; }
</style>
""", unsafe_allow_html=True)


def ler_xml(conteudo: bytes) -> pd.DataFrame:
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError as e:
        raise ValueError(f"XML inválido: {e}")

    NS = "urn:schemas-microsoft-com:office:spreadsheet"
    def tag(nome): return f"{{{NS}}}{nome}"

    worksheet = root.find(f".//{tag('Worksheet')}")
    if worksheet is not None:
        table = worksheet.find(tag("Table"))
        if table is None:
            for ws in root.iter(tag("Worksheet")):
                table = ws.find(tag("Table"))
                if table is not None: break
        if table is not None:
            linhas = list(table.findall(tag("Row")))
            if not linhas: raise ValueError("SpreadsheetML: nenhuma linha encontrada na tabela.")
            ATTR_INDEX = f"{{{NS}}}Index"

            def celulas(row_el):
                vals = []; idx = 0
                for cell in row_el.findall(tag("Cell")):
                    ss_idx = cell.get(ATTR_INDEX)
                    if ss_idx:
                        alvo = int(ss_idx) - 1
                        while idx < alvo: vals.append(""); idx += 1
                    data_el = cell.find(tag("Data"))
                    vals.append((data_el.text or "").strip() if data_el is not None else "")
                    idx += 1
                return vals

            cabecalho = celulas(linhas[0])
            cab_norm = []; contagem = {}
            for c in cabecalho:
                c = c.strip() if c.strip() else f"_col{len(cab_norm)}"
                contagem[c] = contagem.get(c, 0) + 1
                cab_norm.append(c if contagem[c] == 1 else f"{c}_{contagem[c]}")

            registros = []
            for row_el in list(linhas)[1:]:
                vals = celulas(row_el)
                while len(vals) < len(cab_norm): vals.append("")
                registros.append(dict(zip(cab_norm, vals[:len(cab_norm)])))

            if not registros: raise ValueError("SpreadsheetML: tabela sem linhas de dados.")
            df = pd.DataFrame(registros)
            df.columns = [str(c).strip() for c in df.columns]
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)].reset_index(drop=True)
            return df
    return pd.DataFrame()


def ler_arquivo(arquivo) -> pd.DataFrame:
    nome = arquivo.name.lower(); conteudo = arquivo.read()
    if nome.endswith(".xml"): return ler_xml(conteudo)
    elif nome.endswith((".xlsx", ".xls")): return pd.read_excel(io.BytesIO(conteudo))
    else: raise ValueError("Formato de arquivo não suportado.")


def preparar_dados(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip()
    if 'Situação' in df.columns:
        df['Situação'] = df['Situação'].astype(str).str.strip()
        df = df[df['Situação'] == 'Atribuida'].reset_index(drop=True)
    else:
        st.warning("⚠️ Coluna 'Situação' não encontrada. Nenhum filtro de situação foi aplicado.")

    if 'Valor Faturas' in df.columns:
        def _parse_valor(v):
            s = str(v).strip().replace('R$', '').replace(' ', '')
            if not s or s in ('nan', 'None', '-'): return 0.0
            if ',' in s: s = s.replace('.', '').replace(',', '.')
            try: return float(s)
            except ValueError: return 0.0
        df['Valor Faturas'] = df['Valor Faturas'].apply(_parse_valor)

    if 'Quantidade Faturas' in df.columns:
        df['Quantidade Faturas'] = pd.to_numeric(df['Quantidade Faturas'], errors='coerce').fillna(0).astype(int)

    if 'Data Inclusão' in df.columns:
        df['Data Inclusão'] = pd.to_datetime(df['Data Inclusão'], dayfirst=True, errors='coerce')

    return df


# ─── COLUNAS DE SAÍDA — Endereço e Bairro incluídos ───
COLUNAS_SAIDA = [
    'Numero', 'Subtipo', 'Data Inclusão', 'Prefixo',
    'Instalação CCS', 'Situação', 'Valor Faturas', 'Quantidade Faturas', 'Endereço', 'Bairro'
]

SUBTIPOS_EXCLUIR_ARRECADADO = [
    "P2 VISTORIA - GRUPO A", "P1 SUSPENSÃO - GRUPO A", "P3 SUSPENSÃO - GRUPO A",
    "P2 SUSPENSÃO - GRUPO A", "P3 VISTORIA - GRUPO A", "P1 VISTORIA - GRUPO A",
]
SUBTIPOS_SUSPENSAO_P1 = [
    "P1 SUSPENSÃO - GRUPO A", "P1 SUSPENSÃO - POSTE", "P1 VISTORIA - RETIRADA DE RAMAL",
]
SUBTIPOS_GRUPO_A = [
    "P2 VISTORIA - GRUPO A", "P1 SUSPENSÃO - GRUPO A", "P3 SUSPENSÃO - GRUPO A",
    "P2 SUSPENSÃO - GRUPO A", "P3 VISTORIA - GRUPO A", "P1 VISTORIA - GRUPO A",
]


def selecionar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in COLUNAS_SAIDA if c in df.columns]
    ausentes = [c for c in COLUNAS_SAIDA if c not in df.columns]
    if ausentes:
        st.warning(f"⚠️ Colunas não encontradas no arquivo e omitidas: {', '.join(ausentes)}")
    return df[cols]


def modalidade_arrecadado(df, limite_faturas, limite_prefixo):
    result = df.copy()
    if 'Subtipo' in result.columns: result = result[~result['Subtipo'].isin(SUBTIPOS_EXCLUIR_ARRECADADO)]
    if 'Quantidade Faturas' in result.columns: result = result[result['Quantidade Faturas'] <= limite_faturas]
    if 'Valor Faturas' in result.columns: result = result.sort_values('Valor Faturas', ascending=False)
    if 'Prefixo' in result.columns: result = result.groupby('Prefixo', group_keys=False).head(limite_prefixo)
    return selecionar_colunas(result)


def modalidade_quantidade_faturas(df, piso_faturas, limite_prefixo):
    result = df.copy()
    if 'Quantidade Faturas' in result.columns:
        result = result[result['Quantidade Faturas'] >= piso_faturas]
        result = result.sort_values('Quantidade Faturas', ascending=False)
    if 'Prefixo' in result.columns: result = result.groupby('Prefixo', group_keys=False).head(limite_prefixo)
    return selecionar_colunas(result)


def modalidade_suspensao_p1(df, limite_prefixo):
    result = df.copy()
    if 'Subtipo' in result.columns: result = result[result['Subtipo'].isin(SUBTIPOS_SUSPENSAO_P1)]
    sort_cols, ascendings = [], []
    if 'Data Inclusão' in result.columns: sort_cols.append('Data Inclusão'); ascendings.append(True)
    if 'Valor Faturas' in result.columns: sort_cols.append('Valor Faturas'); ascendings.append(False)
    if sort_cols: result = result.sort_values(sort_cols, ascending=ascendings)
    if 'Prefixo' in result.columns: result = result.groupby('Prefixo', group_keys=False).head(limite_prefixo)
    return selecionar_colunas(result)


def modalidade_grupo_a(df, limite_prefixo):
    result = df.copy()
    if 'Subtipo' in result.columns: result = result[result['Subtipo'].isin(SUBTIPOS_GRUPO_A)]
    if 'Valor Faturas' in result.columns: result = result.sort_values('Valor Faturas', ascending=False)
    if 'Prefixo' in result.columns: result = result.groupby('Prefixo', group_keys=False).head(limite_prefixo)
    return selecionar_colunas(result)


def gerar_excel(df: pd.DataFrame, modalidade: str) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = modalidade[:30]
    COR_HEADER_BG = "0D2B6B"; COR_HEADER_FG = "FFFFFF"
    COR_LINHA_PAR = "EEF3FB"; COR_LINHA_IMPAR = "FFFFFF"; COR_BORDA = "C7D6ED"

    header_fill  = PatternFill("solid", fgColor=COR_HEADER_BG)
    header_font  = Font(bold=True, color=COR_HEADER_FG, size=10, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    par_fill     = PatternFill("solid", fgColor=COR_LINHA_PAR)
    impar_fill   = PatternFill("solid", fgColor=COR_LINHA_IMPAR)
    thin  = Side(style="thin", color=COR_BORDA)
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(size=9, name="Calibri")

    headers = list(df.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = header_align; cell.border = borda
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = par_fill if row_idx % 2 == 0 else impar_fill
        for col_idx, (col_name, value) in enumerate(zip(headers, row), start=1):
            if col_name == 'Data Inclusão' and pd.notna(value):
                try: cell_val = pd.Timestamp(value).to_pydatetime()
                except Exception: cell_val = value
            else:
                cell_val = value
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.fill = fill; cell.font = body_font; cell.border = borda
            if col_name == 'Valor Faturas':
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = 'R$ #,##0.00'
            elif col_name == 'Quantidade Faturas':
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == 'Data Inclusão':
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = 'DD/MM/YYYY'
            elif col_name in ('Numero', 'Prefixo', 'Instalação CCS', 'Situação'):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ('Endereço', 'Bairro'):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        # Endereço e Bairro podem ser longos — limita um pouco mais
        limite = 60 if col_name == 'Endereço' else 35 if col_name == 'Bairro' else 45
        ws.column_dimensions[col_letter].width = min(max_len + 4, limite)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO(); wb.save(buffer); return buffer.getvalue()


# ─── INTERFACE ───
st.markdown("""
<div class="main-header">
    <h1>⚡ Relatório de Serviços Prioritários</h1>
    <p>Automatize a geração do relatório diário com filtros inteligentes por modalidade.</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown('<p class="sidebar-title">📁 Entrada de Dados</p>', unsafe_allow_html=True)
    arquivo = st.file_uploader("Carregar arquivo", type=["xlsx", "xls", "xml"],
        help="Formatos aceitos: .xlsx, .xls, .xml (SpreadsheetML)")
    st.markdown("---")
    st.markdown('<p class="sidebar-title">🎯 Configuração</p>', unsafe_allow_html=True)
    modalidade = st.selectbox("Modalidade de Prioridade",
        options=["ARRECADADO", "QUANTIDADE DE FATURAS", "SUSPENSÃO E VISTORIA P1", "GRUPO A"],
        help="Selecione a lógica de filtragem e ordenação desejada.")
    st.markdown("---")
    st.markdown('<p class="sidebar-title">⚙️ Parâmetros</p>', unsafe_allow_html=True)
    limite_prefixo = st.number_input("Máx. de serviços por Prefixo", min_value=1, max_value=100, value=5, step=1)
    limite_faturas = None; piso_faturas = None

    if modalidade == "ARRECADADO":
        st.markdown('<div class="info-box">🔹 Ordena por <b>Valor Faturas</b> (maior → menor). Exclui subtipos de Grupo A e Vistoria.</div>', unsafe_allow_html=True)
        limite_faturas = st.number_input("Limite máximo de Qtd. Faturas", min_value=1, max_value=99999, value=10, step=1)
    elif modalidade == "QUANTIDADE DE FATURAS":
        st.markdown('<div class="info-box">🔹 Ordena por <b>Quantidade de Faturas</b> (maior → menor).</div>', unsafe_allow_html=True)
        teto_faturas = st.number_input("Teto máximo de Qtd. Faturas", min_value=1, max_value=99999, value=10, step=1)
    elif modalidade == "SUSPENSÃO E VISTORIA P1":
        st.markdown('<div class="info-box">🔹 Filtra subtipos P1. Ordena por <b>Data Inclusão</b> (mais antiga) e desempata por <b>Valor Faturas</b> (maior).</div>', unsafe_allow_html=True)
    elif modalidade == "GRUPO A":
        st.markdown('<div class="info-box">🔹 Filtra subtipos Grupo A e ordena por <b>Valor Faturas</b> (maior → menor).</div>', unsafe_allow_html=True)

if arquivo is None:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div style="text-align:center; padding: 3rem 1rem; color: #94a3b8;">
            <div style="font-size: 3rem; margin-bottom: 1rem;">📂</div>
            <div style="font-size: 1.1rem; font-weight: 600; color: #64748b;">Nenhum arquivo carregado</div>
            <div style="font-size: 0.9rem; margin-top: 0.5rem;">
                Use o painel lateral para carregar um arquivo <b>.xlsx</b>, <b>.xls</b> ou <b>.xml</b>.
            </div>
        </div>
        """, unsafe_allow_html=True)
else:
    try:
        with st.spinner("📊 Lendo e processando o arquivo..."):
            df_raw = ler_arquivo(arquivo)
            df_clean = preparar_dados(df_raw)

        total_raw = len(df_raw); total_clean = len(df_clean)
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f'<div class="metric-card"><div class="label">Total de Registros</div><div class="value">{total_raw:,}</div></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="metric-card"><div class="label">Registros "Atribuida"</div><div class="value">{total_clean:,}</div></div>', unsafe_allow_html=True)
        with col3:
            perc = round((total_clean / total_raw * 100) if total_raw else 0, 1)
            st.markdown(f'<div class="metric-card"><div class="label">Taxa de Aproveitamento</div><div class="value">{perc}%</div></div>', unsafe_allow_html=True)
        with col4:
            prefixos = df_clean['Prefixo'].nunique() if 'Prefixo' in df_clean.columns else 0
            st.markdown(f'<div class="metric-card"><div class="label">Prefixos Distintos</div><div class="value">{prefixos}</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        with st.spinner(f"⚙️ Aplicando filtros da modalidade **{modalidade}**..."):
            if modalidade == "ARRECADADO":
                df_result = modalidade_arrecadado(df_clean, limite_faturas, limite_prefixo)
            elif modalidade == "QUANTIDADE DE FATURAS":
                df_result = modalidade_quantidade_faturas(df_clean, teto_faturas, limite_prefixo)
            elif modalidade == "SUSPENSÃO E VISTORIA P1":
                df_result = modalidade_suspensao_p1(df_clean, limite_prefixo)
            elif modalidade == "GRUPO A":
                df_result = modalidade_grupo_a(df_clean, limite_prefixo)

        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:0.75rem;">
            <span class="badge">✅ {modalidade}</span>
            <span style="color:#64748b; font-size:0.9rem;">{len(df_result):,} registros encontrados</span>
        </div>
        """, unsafe_allow_html=True)

        if df_result.empty:
            st.warning("⚠️ Nenhum registro encontrado com os filtros aplicados. Tente ajustar os parâmetros.")
        else:
            st.dataframe(df_result, use_container_width=True, height=420, hide_index=True)
            st.markdown("<br>", unsafe_allow_html=True)
            excel_bytes = gerar_excel(df_result, modalidade)
            nome_arquivo = f"relatorio_{modalidade.lower().replace(' ', '_')}.xlsx"
            st.download_button(
                label=f"⬇️  Baixar Relatório Excel — {modalidade}",
                data=excel_bytes, file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except ValueError as e:
        st.error(f"❌ Erro ao processar o arquivo: {e}")
    except Exception as e:
        st.error(f"❌ Erro inesperado: {e}"); raise e
