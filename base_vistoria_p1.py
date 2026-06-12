import streamlit as st
import pandas as pd
import io
import requests
import xml.etree.ElementTree as ET
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Credenciais WhatsApp Cloud API ───────────────────────────────────────────
WA_TOKEN     = st.secrets.get("WHATSAPP_TOKEN", "")
WA_PHONE_ID  = st.secrets.get("PHONE_NUMBER_ID", "")
WA_RECIPIENT = st.secrets.get("RECIPIENT_NUMBER", "")


# ── Mapeamento de nomes alternativos de colunas ───────────────────────────────
# Chave = nome canônico usado no código | Valor = lista de aliases aceitos
COLUMN_ALIASES = {
    "Subtipo":        ["Subtipo", "SubTipoDescricao"],
    "Data Inclusão":  ["Data Inclusão", "Data Inclusao", "DataInclusao"],
    "Numero":         ["Numero"],
    "Valor Faturas":  ["Valor Faturas", "ValorFaturasUc"],
}

def normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    """Renomeia colunas usando os aliases definidos em COLUMN_ALIASES."""
    renomear = {}
    for canonico, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in df.columns and alias != canonico:
                renomear[alias] = canonico
    return df.rename(columns=renomear)


# ── Leitura de XML ────────────────────────────────────────────────────────────

def ler_xml(conteudo: bytes) -> pd.DataFrame:
    """
    Lê XML e retorna DataFrame.

    Suporta três formatos:
      1. SpreadsheetML — XML gerado pelo Excel/Office com namespace
         urn:schemas-microsoft-com:office:spreadsheet  (ex.: exportações do Oper)
      2. XML genérico com sub-elementos como colunas
         <root><row><Col1>val</Col1></row></root>
      3. XML genérico com atributos como colunas
         <root><row Col1="val"/></root>
    """
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError as e:
        raise ValueError(f"XML inválido: {e}")

    # ── 1. SpreadsheetML (namespace Office/Excel) ──────────────────────────
    NS = "urn:schemas-microsoft-com:office:spreadsheet"

    def tag(nome):
        return f"{{{NS}}}{nome}"

    worksheet = root.find(f".//{tag('Worksheet')}")
    if worksheet is not None:
        table = worksheet.find(tag("Table"))
        if table is None:
            for ws in root.iter(tag("Worksheet")):
                table = ws.find(tag("Table"))
                if table is not None:
                    break

        if table is not None:
            linhas = list(table.findall(tag("Row")))
            if not linhas:
                raise ValueError("SpreadsheetML: nenhuma linha encontrada na tabela.")

            ATTR_INDEX = f"{{{NS}}}Index"

            def celulas(row_el):
                vals = []
                idx = 0
                for cell in row_el.findall(tag("Cell")):
                    ss_idx = cell.get(ATTR_INDEX)
                    if ss_idx:
                        alvo = int(ss_idx) - 1
                        while idx < alvo:
                            vals.append("")
                            idx += 1
                    data_el = cell.find(tag("Data"))
                    vals.append((data_el.text or "").strip() if data_el is not None else "")
                    idx += 1
                return vals

            cabecalho = celulas(linhas[0])
            cab_norm = []
            contagem: dict = {}
            for c in cabecalho:
                c = c.strip() if c.strip() else f"_col{len(cab_norm)}"
                contagem[c] = contagem.get(c, 0) + 1
                cab_norm.append(c if contagem[c] == 1 else f"{c}_{contagem[c]}")

            registros = []
            for row_el in linhas[1:]:
                vals = celulas(row_el)
                while len(vals) < len(cab_norm):
                    vals.append("")
                registros.append(dict(zip(cab_norm, vals[:len(cab_norm)])))

            if not registros:
                raise ValueError("SpreadsheetML: tabela sem linhas de dados.")

            df = pd.DataFrame(registros)
            df.columns = [str(c).strip() for c in df.columns]
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)].reset_index(drop=True)
            return df

    # ── 2 & 3. XML genérico ────────────────────────────────────────────────
    filhos = list(root)
    if not filhos:
        raise ValueError("XML sem elementos filhos na raiz.")

    candidatos = filhos
    for _ in range(3):
        netos = []
        for f in candidatos:
            netos.extend(list(f))
        if len(netos) > len(candidatos):
            candidatos = netos
        else:
            break

    amostra = candidatos[0]
    registros = []

    if list(amostra):
        def limpar_tag(t):
            return t.split("}")[-1] if "}" in t else t

        for filho in candidatos:
            registro = {limpar_tag(sub.tag): (sub.text or "").strip() for sub in filho}
            if registro:
                registros.append(registro)
    elif amostra.attrib:
        for filho in candidatos:
            if filho.attrib:
                registros.append(dict(filho.attrib))

    if not registros:
        raise ValueError(
            "Não foi possível extrair registros do XML. "
            "Verifique se o arquivo possui dados tabulares."
        )

    df = pd.DataFrame(registros)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ── Leitura de arquivo (xlsx, xls, xml) ──────────────────────────────────────

def ler_arquivo(arquivo) -> pd.DataFrame:
    """Lê xlsx, xls ou xml e retorna DataFrame com colunas normalizadas."""
    nome = arquivo.name.lower()
    conteudo = arquivo.read()
    arquivo.seek(0)

    if nome.endswith(".xml"):
        df = ler_xml(conteudo)
    elif nome.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(conteudo), sheet_name=0)
    else:
        raise ValueError(f"Formato de arquivo não suportado: {nome}")

    return normalizar_colunas(df)


# ── Gerar Excel ───────────────────────────────────────────────────────────────

def gerar_excel(tabela: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Acomp. P1"

    AZUL_HEADER = "2E5F9E"
    AZUL_TITULO = "1F4E79"
    BRANCO      = "FFFFFF"
    CINZA_LINHA = "DEEAF1"

    thin = Side(style="thin", color="AAAAAA")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Acompanhamento de Suspensão/Vistoria P1"
    c.font = Font(name="Arial", bold=True, color=BRANCO, size=13)
    c.fill = PatternFill("solid", fgColor=AZUL_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["DATA", "P1 SUSPENSÃO - GRUPO A", "P1 SUSPENSÃO - POSTE",
               "P1 VISTORIA - RETIRADA DE RAMAL", "Total Geral", "Total Dívida"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = Font(name="Arial", bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
    ws.row_dimensions[2].height = 30

    meses = {"Jan":"jan","Feb":"fev","Mar":"mar","Apr":"abr","May":"mai","Jun":"jun",
             "Jul":"jul","Aug":"ago","Sep":"set","Oct":"out","Nov":"nov","Dec":"dez"}

    for i, row in enumerate(tabela.itertuples(index=False), start=3):
        fill_cor = PatternFill("solid", fgColor=CINZA_LINHA if i % 2 == 0 else BRANCO)
        data_str = row[0].strftime("%d/%b")
        for en, pt in meses.items():
            data_str = data_str.replace(en, pt)

        c = ws.cell(row=i, column=1, value=data_str)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        c.fill = fill_cor
        c.border = borda

        for col_idx, val in enumerate([row[1], row[2], row[3], row[4]], start=2):
            c = ws.cell(row=i, column=col_idx, value=int(val))
            c.font = Font(name="Arial", size=10, bold=(col_idx == 5))
            c.alignment = Alignment(horizontal="center")
            c.fill = fill_cor
            c.border = borda

        c = ws.cell(row=i, column=6, value=row[5])
        c.number_format = 'R$ #,##0.00'
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="right")
        c.fill = fill_cor
        c.border = borda

    total_row = ws.max_row + 1
    data_start, data_end = 3, total_row - 1
    ws.cell(row=total_row, column=1, value="Total Geral").font = Font(name="Arial", bold=True, color=BRANCO, size=10)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=AZUL_HEADER)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=1).border = borda

    for col_idx, col_letter in enumerate(["B","C","D","E","F"], start=2):
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        c.font = Font(name="Arial", bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal="center" if col_idx < 6 else "right")
        c.border = borda
        if col_idx == 6:
            c.number_format = 'R$ #,##0.00'

    for i, w in enumerate([12, 26, 22, 34, 14, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Gerar Imagem da Tabela ────────────────────────────────────────────────────

def gerar_imagem_tabela(tabela: pd.DataFrame, divida_dia: pd.Series) -> bytes:
    meses = {"Jan":"jan","Feb":"fev","Mar":"mar","Apr":"abr","May":"mai","Jun":"jun",
             "Jul":"jul","Aug":"ago","Sep":"set","Oct":"out","Nov":"nov","Dec":"dez"}

    linhas = []
    for row in tabela.itertuples(index=False):
        data_str = row[0].strftime("%d/%b")
        for en, pt in meses.items():
            data_str = data_str.replace(en, pt)
        divida_fmt = f"R$ {float(row[5]):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        linhas.append([
            data_str,
            str(int(row[1])),
            str(int(row[2])),
            str(int(row[3])),
            str(int(row[4])),
            divida_fmt,
        ])

    t1 = int(tabela.iloc[:,1].sum())
    t2 = int(tabela.iloc[:,2].sum())
    t3 = int(tabela.iloc[:,3].sum())
    t4 = int(tabela.iloc[:,4].sum())
    t5 = divida_dia.sum()
    t5_fmt = f"R$ {t5:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    linhas.append(["TOTAL GERAL", str(t1), str(t2), str(t3), str(t4), t5_fmt])

    headers = ["DATA", "P1 SUSPENSÃO\nGRUPO A", "P1 SUSPENSÃO\nPOSTE",
               "P1 VISTORIA\nRET. RAMAL", "TOTAL\nGERAL", "TOTAL DÍVIDA"]

    n_rows = len(linhas)
    n_cols = len(headers)

    COR_TITULO  = "#1F4E79"
    COR_HEADER  = "#2E5F9E"
    COR_LISTA   = "#DEEAF1"
    COR_BRANCO  = "#FFFFFF"
    COR_TOTAL   = "#2E5F9E"
    COR_TEXTO   = "#1a1a2e"
    COR_TXBCO   = "#FFFFFF"

    fig_w = 16
    row_h = 0.55
    header_h = 0.75
    title_h = 0.65
    fig_h = title_h + header_h + n_rows * row_h + 0.3

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    fig.patch.set_facecolor(COR_BRANCO)
    ax.set_facecolor(COR_BRANCO)
    ax.axis("off")

    total_h = fig_h
    y_cursor = total_h

    y_cursor -= title_h
    title_rect = mpatches.FancyBboxPatch(
        (0, y_cursor), fig_w, title_h,
        boxstyle="square,pad=0",
        linewidth=0, facecolor=COR_TITULO,
        transform=ax.transData, clip_on=False
    )
    ax.add_patch(title_rect)
    ax.text(fig_w / 2, y_cursor + title_h / 2,
            "Acompanhamento de Suspensão / Vistoria P1",
            ha="center", va="center",
            fontsize=13, fontweight="bold", color=COR_TXBCO,
            transform=ax.transData)

    y_cursor -= header_h
    col_widths = [1.6, 2.8, 2.4, 2.8, 1.6, 2.4]
    total_w = sum(col_widths)
    scale = fig_w / total_w
    col_widths_px = [w * scale for w in col_widths]

    x = 0
    for j, (h, cw) in enumerate(zip(headers, col_widths_px)):
        rect = mpatches.FancyBboxPatch(
            (x, y_cursor), cw, header_h,
            boxstyle="square,pad=0",
            linewidth=0, facecolor=COR_HEADER,
            transform=ax.transData, clip_on=False
        )
        ax.add_patch(rect)
        ax.text(x + cw / 2, y_cursor + header_h / 2, h,
                ha="center", va="center",
                fontsize=8.5, fontweight="bold", color=COR_TXBCO,
                linespacing=1.3, transform=ax.transData)
        x += cw

    for i, linha in enumerate(linhas):
        y_cursor -= row_h
        is_total = (i == len(linhas) - 1)
        cor_fundo = COR_TOTAL if is_total else (COR_LISTA if i % 2 == 0 else COR_BRANCO)
        cor_txt   = COR_TXBCO if is_total else COR_TEXTO
        peso      = "bold" if is_total else "normal"

        x = 0
        for j, (val, cw) in enumerate(zip(linha, col_widths_px)):
            rect = mpatches.FancyBboxPatch(
                (x, y_cursor), cw, row_h,
                boxstyle="square,pad=0",
                linewidth=0, facecolor=cor_fundo,
                transform=ax.transData, clip_on=False
            )
            ax.add_patch(rect)

            align = "right" if j == n_cols - 1 else "center"
            x_txt = (x + cw - 0.15) if align == "right" else (x + cw / 2)
            ax.text(x_txt, y_cursor + row_h / 2, val,
                    ha=align, va="center",
                    fontsize=9, fontweight=peso, color=cor_txt,
                    transform=ax.transData)
            x += cw

    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, total_h)

    plt.tight_layout(pad=0)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=COR_BRANCO)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ── WhatsApp Cloud API ────────────────────────────────────────────────────────

def upload_midia_whatsapp(img_bytes: bytes) -> str | None:
    url = f"https://graph.facebook.com/v19.0/{WA_PHONE_ID}/media"
    headers = {"Authorization": f"Bearer {WA_TOKEN}"}
    files = {
        "file": ("relatorio_p1.png", img_bytes, "image/png"),
        "type": (None, "image/png"),
        "messaging_product": (None, "whatsapp"),
    }
    resp = requests.post(url, headers=headers, files=files)
    if resp.status_code == 200:
        return resp.json().get("id")
    st.error(f"Erro no upload da mídia: {resp.status_code} — {resp.text}")
    return None


def enviar_imagem_whatsapp(media_id: str, numero: str) -> bool:
    url = f"https://graph.facebook.com/v19.0/{WA_PHONE_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WA_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "image",
        "image": {
            "id": media_id,
            "caption": "📋 *Acompanhamento Suspensão/Vistoria P1*\nRelatório gerado automaticamente.",
        },
    }
    resp = requests.post(url, headers=headers, json=payload)
    if resp.status_code == 200:
        return True
    st.error(f"Erro ao enviar imagem: {resp.status_code} — {resp.text}")
    return False


def enviar_whatsapp(img_bytes: bytes, numero: str) -> bool:
    if not WA_TOKEN or not WA_PHONE_ID:
        st.error("⚠️ Credenciais do WhatsApp não configuradas nos Secrets do Streamlit.")
        return False
    with st.spinner("📤 Fazendo upload da imagem..."):
        media_id = upload_midia_whatsapp(img_bytes)
    if not media_id:
        return False
    with st.spinner("💬 Enviando pelo WhatsApp..."):
        return enviar_imagem_whatsapp(media_id, numero)


# ── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Acompanhamento Suspensão/Vistoria P1", layout="wide")
st.title("📋 Acompanhamento de Suspensão/Vistoria P1")

uploaded_file = st.file_uploader("Suba a base (.xlsx ou .xml)", type=["xlsx", "xml"])

SUBTIPOS_P1 = [
    "P1 SUSPENSÃO - GRUPO A",
    "P1 SUSPENSÃO - POSTE",
    "P1 VISTORIA - RETIRADA DE RAMAL",
]

if uploaded_file:
    try:
        df = ler_arquivo(uploaded_file)

        required_cols = {"Subtipo", "Data Inclusão", "Numero", "Valor Faturas"}
        if not required_cols.issubset(df.columns):
            faltando = required_cols - set(df.columns)
            st.error(
                f"Colunas esperadas não encontradas após normalização: {faltando}\n\n"
                f"Colunas encontradas no arquivo: {list(df.columns)}"
            )
            st.stop()

        df_p1 = df[df["Subtipo"].isin(SUBTIPOS_P1)].copy()
        df_p1["Data Inclusão"] = pd.to_datetime(df_p1["Data Inclusão"]).dt.date

        pivot_qtd = (
            df_p1.groupby(["Data Inclusão", "Subtipo"])["Numero"]
            .count()
            .unstack(fill_value=0)
            .reindex(columns=SUBTIPOS_P1, fill_value=0)
        )
        pivot_qtd["Total Geral"] = pivot_qtd.sum(axis=1)

        divida_dia = (
            df_p1.groupby("Data Inclusão")["Valor Faturas"]
            .sum()
            .rename("Total Dívida")
        )

        tabela = pivot_qtd.join(divida_dia).reset_index()
        tabela = tabela.rename(columns={"Data Inclusão": "DATA"})
        tabela = tabela.sort_values("DATA")

        totais = tabela.drop(columns="DATA").sum()
        totais_row = pd.DataFrame([["Total Geral"] + totais.tolist()], columns=tabela.columns)
        tabela_exibir = pd.concat([tabela, totais_row], ignore_index=True)

        tabela_exibir["DATA"] = tabela_exibir["DATA"].apply(
            lambda x: x.strftime("%d/%b").replace("/0", "/").replace(
                "Jan","jan").replace("Feb","fev").replace("Mar","mar").replace(
                "Apr","abr").replace("May","mai").replace("Jun","jun").replace(
                "Jul","jul").replace("Aug","ago").replace("Sep","set").replace(
                "Oct","out").replace("Nov","nov").replace("Dec","dez")
            if hasattr(x, "strftime") else str(x)
        )

        tabela_exibir["Total Dívida"] = tabela_exibir["Total Dívida"].apply(
            lambda x: f"R$ {float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            if x != "Total Dívida" else x
        )

        # Métricas
        st.subheader("Resumo do Período")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("P1 Suspensão - Grupo A", int(pivot_qtd["P1 SUSPENSÃO - GRUPO A"].sum()))
        col2.metric("P1 Suspensão - Poste", int(pivot_qtd["P1 SUSPENSÃO - POSTE"].sum()))
        col3.metric("P1 Vistoria - Ret. Ramal", int(pivot_qtd["P1 VISTORIA - RETIRADA DE RAMAL"].sum()))
        total_divida = divida_dia.sum()
        total_fmt = f"R$ {total_divida:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        col4.metric("Total Dívida", total_fmt)

        st.divider()

        excel_bytes = gerar_excel(tabela)
        img_bytes   = gerar_imagem_tabela(tabela, divida_dia)

        st.subheader("Detalhamento por Dia")
        st.image(img_bytes, use_container_width=True)

        st.divider()

        col_excel, col_whats = st.columns([1, 1])

        col_excel.download_button(
            label="⬇️ Baixar Excel",
            data=excel_bytes,
            file_name="acompanhamento_suspensao_p1.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        with col_whats:
            if st.button("💬 Enviar imagem pelo WhatsApp", use_container_width=True):
                st.session_state["mostrar_envio_zap"] = True

        if st.session_state.get("mostrar_envio_zap"):
            with st.expander("📤 Confirmar envio pelo WhatsApp", expanded=True):
                numero = st.text_input(
                    "📱 Número do destinatário",
                    placeholder="Ex: 5562999999999  (DDI + DDD + número, sem espaços)",
                    key="numero_destinatario"
                )
                st.caption("🇧🇷 Brasil: comece com 55 — ex: **5562999887766**")

                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button("✅ Confirmar envio", use_container_width=True):
                        if not numero.strip():
                            st.warning("Digite o número antes de enviar.")
                        elif not numero.strip().isdigit() or len(numero.strip()) < 12:
                            st.warning("Número inválido. Use apenas dígitos com DDI+DDD, ex: 5562999887766")
                        else:
                            sucesso = enviar_whatsapp(img_bytes, numero.strip())
                            if sucesso:
                                st.success("✅ Imagem enviada com sucesso!")
                                st.session_state["mostrar_envio_zap"] = False
                with col_b:
                    if st.button("❌ Cancelar", use_container_width=True):
                        st.session_state["mostrar_envio_zap"] = False
                        st.rerun()

        st.dataframe(
            tabela_exibir,
            use_container_width=True,
            hide_index=True,
            column_config={
                "DATA": st.column_config.TextColumn("DATA"),
                "P1 SUSPENSÃO - GRUPO A": st.column_config.NumberColumn("P1 SUSPENSÃO - GRUPO A"),
                "P1 SUSPENSÃO - POSTE": st.column_config.NumberColumn("P1 SUSPENSÃO - POSTE"),
                "P1 VISTORIA - RETIRADA DE RAMAL": st.column_config.NumberColumn("P1 VISTORIA - RETIRADA DE RAMAL"),
                "Total Geral": st.column_config.NumberColumn("Total Geral"),
                "Total Dívida": st.column_config.TextColumn("Total Dívida"),
            },
        )

    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")
else:
    st.info("👆 Suba um arquivo .xlsx ou .xml para começar.")
