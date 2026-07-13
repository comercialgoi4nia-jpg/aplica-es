import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Relatório de Serviços Prioritários",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main-header {
        background: linear-gradient(135deg, #0d2b6b 0%, #1a52b3 60%, #2176ff 100%);
        border-radius: 14px; padding: 2rem 2.5rem; margin-bottom: 1.5rem;
        color: white; box-shadow: 0 8px 32px rgba(13,43,107,0.25);
    }
    .main-header h1 { font-size: 2rem; font-weight: 700; margin: 0 0 0.3rem 0; letter-spacing: -0.5px; }
    .main-header p { font-size: 0.95rem; opacity: 0.85; margin: 0; }
    .metric-card {
        background: white; border: 1px solid #e2e8f0; border-radius: 10px;
        padding: 1rem 1.25rem; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .metric-card .label { font-size: 0.75rem; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: 0.6px; }
    .metric-card .value { font-size: 1.8rem; font-weight: 700; color: #0d2b6b; line-height: 1.2; }
    .badge { display: inline-block; padding: 0.25rem 0.75rem; border-radius: 20px; font-size: 0.75rem; font-weight: 600; background: #e8f0fe; color: #1a52b3; margin-bottom: 1rem; }
    .info-box { background: #f0f7ff; border-left: 4px solid #2176ff; border-radius: 0 8px 8px 0; padding: 0.85rem 1.1rem; font-size: 0.9rem; color: #1e3a5f; margin-bottom: 1rem; }
    div[data-testid="stSidebar"] { background: #f8fafc; border-right: 1px solid #e2e8f0; }
    div[data-testid="stSidebar"] .sidebar-title { font-size: 0.7rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; color: #94a3b8; padding: 0.5rem 0 0.25rem 0; }
    .stDownloadButton > button {
        width: 100%; background: linear-gradient(135deg, #0d2b6b, #2176ff) !important;
        color: white !important; border: none !important; border-radius: 8px !important;
        font-weight: 600 !important; padding: 0.65rem 1rem !important; font-size: 0.95rem !important;
        transition: opacity 0.2s; box-shadow: 0 4px 12px rgba(33,118,255,0.3) !important;
    }
    .stDownloadButton > button:hover { opacity: 0.9; }
    .stSelectbox label, .stNumberInput label, .stFileUploader label { font-weight: 600 !important; font-size: 0.85rem !important; color: #334155 !important; }
    .modal-card {
        background: white; border: 1px solid #e2e8f0; border-radius: 10px;
        padding: 1rem 1.1rem; box-shadow: 0 2px 8px rgba(0,0,0,0.05); margin-bottom: 0.75rem;
    }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# LEITURA E PREPARAÇÃO DE DADOS
# ══════════════════════════════════════════════════════════════════

def ler_xml(conteudo: bytes) -> pd.DataFrame:
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError as e:
        raise ValueError(f"XML inválido: {e}")

    NS = "urn:schemas-microsoft-com:office:spreadsheet"
    def tag(nome): return f"{{{NS}}}{nome}"

    worksheet = root.find(f".//{tag('Worksheet')}")
    if worksheet is not None:
        table = worksheet.find(tag("Table"))
        if table is None:
            for ws in root.iter(tag("Worksheet")):
                table = ws.find(tag("Table"))
                if table is not None: break
        if table is not None:
            linhas = list(table.findall(tag("Row")))
            if not linhas: raise ValueError("SpreadsheetML: nenhuma linha encontrada na tabela.")
            ATTR_INDEX = f"{{{NS}}}Index"

            def celulas(row_el):
                vals = []; idx = 0
                for cell in row_el.findall(tag("Cell")):
                    ss_idx = cell.get(ATTR_INDEX)
                    if ss_idx:
                        alvo = int(ss_idx) - 1
                        while idx < alvo: vals.append(""); idx += 1
                    data_el = cell.find(tag("Data"))
                    vals.append((data_el.text or "").strip() if data_el is not None else "")
                    idx += 1
                return vals

            cabecalho = celulas(linhas[0])
            cab_norm = []; contagem = {}
            for c in cabecalho:
                c = c.strip() if c.strip() else f"_col{len(cab_norm)}"
                contagem[c] = contagem.get(c, 0) + 1
                cab_norm.append(c if contagem[c] == 1 else f"{c}_{contagem[c]}")

            registros = []
            for row_el in list(linhas)[1:]:
                vals = celulas(row_el)
                while len(vals) < len(cab_norm): vals.append("")
                registros.append(dict(zip(cab_norm, vals[:len(cab_norm)])))

            if not registros: raise ValueError("SpreadsheetML: tabela sem linhas de dados.")
            df = pd.DataFrame(registros)
            df.columns = [str(c).strip() for c in df.columns]
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)].reset_index(drop=True)
            return df
    return pd.DataFrame()


def ler_arquivo(arquivo) -> pd.DataFrame:
    nome = arquivo.name.lower(); conteudo = arquivo.read()
    if nome.endswith(".xml"): return ler_xml(conteudo)
    elif nome.endswith((".xlsx", ".xls")): return pd.read_excel(io.BytesIO(conteudo))
    else: raise ValueError("Formato de arquivo não suportado.")


def preparar_dados(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip()
    if 'Situação' in df.columns:
        df['Situação'] = df['Situação'].astype(str).str.strip()
        df = df[df['Situação'] == 'Atribuida'].reset_index(drop=True)
    else:
        st.warning("⚠️ Coluna 'Situação' não encontrada. Nenhum filtro de situação foi aplicado.")

    if 'Valor Faturas' in df.columns:
        def _parse_valor(v):
            s = str(v).strip().replace('R$', '').replace(' ', '')
            if not s or s in ('nan', 'None', '-'): return 0.0
            if ',' in s: s = s.replace('.', '').replace(',', '.')
            try: return float(s)
            except ValueError: return 0.0
        df['Valor Faturas'] = df['Valor Faturas'].apply(_parse_valor)

    if 'Quantidade Faturas' in df.columns:
        df['Quantidade Faturas'] = pd.to_numeric(df['Quantidade Faturas'], errors='coerce').fillna(0).astype(int)

    if 'Data Inclusão' in df.columns:
        df['Data Inclusão'] = pd.to_datetime(df['Data Inclusão'], dayfirst=True, errors='coerce')

    return df


# ─── COLUNAS DE SAÍDA — Endereço e Bairro incluídos ───
COLUNAS_SAIDA = [
    'Numero', 'Subtipo', 'Data Inclusão', 'Prefixo',
    'Instalação CCS', 'Situação', 'Valor Faturas', 'Quantidade Faturas', 'Endereço', 'Bairro'
]

MODALIDADES = ["ARRECADADO", "QUANTIDADE DE FATURAS", "SUSPENSÃO E VISTORIA P1", "GRUPO A"]

# Coluna usada para ordenar/definir o "corte" de cada modalidade, e direção da ordenação
COLUNA_ORDENACAO = {
    "ARRECADADO": ("Valor Faturas", False),
    "QUANTIDADE DE FATURAS": ("Quantidade Faturas", False),
    "SUSPENSÃO E VISTORIA P1": ("Data Inclusão", True),
    "GRUPO A": ("Valor Faturas", False),
}

SUBTIPOS_EXCLUIR_ARRECADADO = [
    "P2 VISTORIA - GRUPO A", "P1 SUSPENSÃO - GRUPO A", "P3 SUSPENSÃO - GRUPO A",
    "P2 SUSPENSÃO - GRUPO A", "P3 VISTORIA - GRUPO A", "P1 VISTORIA - GRUPO A",
]
SUBTIPOS_SUSPENSAO_P1 = [
    "P1 SUSPENSÃO - GRUPO A", "P1 SUSPENSÃO - POSTE", "P1 VISTORIA - RETIRADA DE RAMAL",
]
SUBTIPOS_GRUPO_A = [
    "P2 VISTORIA - GRUPO A", "P1 SUSPENSÃO - GRUPO A", "P3 SUSPENSÃO - GRUPO A",
    "P2 SUSPENSÃO - GRUPO A", "P3 VISTORIA - GRUPO A", "P1 VISTORIA - GRUPO A",
]


def selecionar_colunas(df: pd.DataFrame, extras=None) -> pd.DataFrame:
    colunas_desejadas = COLUNAS_SAIDA + (extras or [])
    cols = [c for c in colunas_desejadas if c in df.columns]
    return df[cols]


# ══════════════════════════════════════════════════════════════════
# LÓGICA DE MODALIDADE — separada em duas camadas:
#   1) obter_base_modalidade -> aplica filtros + ordenação (SEM limitar por prefixo)
#   2) modalidade_* -> aplica também o limite por prefixo (usado na aba "Relatório")
# A separação permite reaproveitar a base ordenada nas abas de Mesclagem e Cortes.
# ══════════════════════════════════════════════════════════════════

def obter_base_modalidade(df: pd.DataFrame, modalidade: str, limite_faturas=None, teto_faturas=None) -> pd.DataFrame:
    result = df.copy()

    if modalidade == "ARRECADADO":
        if 'Subtipo' in result.columns:
            result = result[~result['Subtipo'].isin(SUBTIPOS_EXCLUIR_ARRECADADO)]
        if limite_faturas is not None and 'Quantidade Faturas' in result.columns:
            result = result[result['Quantidade Faturas'] <= limite_faturas]
        if 'Valor Faturas' in result.columns:
            result = result.sort_values('Valor Faturas', ascending=False)

    elif modalidade == "QUANTIDADE DE FATURAS":
        if teto_faturas is not None and 'Quantidade Faturas' in result.columns:
            result = result[result['Quantidade Faturas'] >= teto_faturas]
        if 'Quantidade Faturas' in result.columns:
            result = result.sort_values('Quantidade Faturas', ascending=False)

    elif modalidade == "SUSPENSÃO E VISTORIA P1":
        if 'Subtipo' in result.columns:
            result = result[result['Subtipo'].isin(SUBTIPOS_SUSPENSAO_P1)]
        sort_cols, ascendings = [], []
        if 'Data Inclusão' in result.columns: sort_cols.append('Data Inclusão'); ascendings.append(True)
        if 'Valor Faturas' in result.columns: sort_cols.append('Valor Faturas'); ascendings.append(False)
        if sort_cols: result = result.sort_values(sort_cols, ascending=ascendings)

    elif modalidade == "GRUPO A":
        if 'Subtipo' in result.columns:
            result = result[result['Subtipo'].isin(SUBTIPOS_GRUPO_A)]
        if 'Valor Faturas' in result.columns:
            result = result.sort_values('Valor Faturas', ascending=False)

    return result


def aplicar_limite_prefixo(df: pd.DataFrame, limite_prefixo: int) -> pd.DataFrame:
    if 'Prefixo' in df.columns:
        return df.groupby('Prefixo', group_keys=False).head(limite_prefixo)
    return df


def modalidade_arrecadado(df, limite_faturas, limite_prefixo):
    result = aplicar_limite_prefixo(obter_base_modalidade(df, "ARRECADADO", limite_faturas=limite_faturas), limite_prefixo)
    return selecionar_colunas(result)


def modalidade_quantidade_faturas(df, piso_faturas, limite_prefixo):
    result = aplicar_limite_prefixo(obter_base_modalidade(df, "QUANTIDADE DE FATURAS", teto_faturas=piso_faturas), limite_prefixo)
    return selecionar_colunas(result)


def modalidade_suspensao_p1(df, limite_prefixo):
    result = aplicar_limite_prefixo(obter_base_modalidade(df, "SUSPENSÃO E VISTORIA P1"), limite_prefixo)
    return selecionar_colunas(result)


def modalidade_grupo_a(df, limite_prefixo):
    result = aplicar_limite_prefixo(obter_base_modalidade(df, "GRUPO A"), limite_prefixo)
    return selecionar_colunas(result)


# ══════════════════════════════════════════════════════════════════
# GERAÇÃO DE EXCEL
# ══════════════════════════════════════════════════════════════════

def gerar_excel(df: pd.DataFrame, nome_aba: str) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = nome_aba[:30]
    COR_HEADER_BG = "0D2B6B"; COR_HEADER_FG = "FFFFFF"
    COR_LINHA_PAR = "EEF3FB"; COR_LINHA_IMPAR = "FFFFFF"; COR_BORDA = "C7D6ED"

    header_fill  = PatternFill("solid", fgColor=COR_HEADER_BG)
    header_font  = Font(bold=True, color=COR_HEADER_FG, size=10, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    par_fill     = PatternFill("solid", fgColor=COR_LINHA_PAR)
    impar_fill   = PatternFill("solid", fgColor=COR_LINHA_IMPAR)
    thin  = Side(style="thin", color=COR_BORDA)
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(size=9, name="Calibri")

    headers = list(df.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = header_align; cell.border = borda
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = par_fill if row_idx % 2 == 0 else impar_fill
        for col_idx, (col_name, value) in enumerate(zip(headers, row), start=1):
            if col_name == 'Data Inclusão' and pd.notna(value):
                try: cell_val = pd.Timestamp(value).to_pydatetime()
                except Exception: cell_val = value
            else:
                cell_val = value
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.fill = fill; cell.font = body_font; cell.border = borda
            if col_name == 'Valor Faturas':
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = 'R$ #,##0.00'
            elif col_name == 'Quantidade Faturas':
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == 'Data Inclusão':
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = 'DD/MM/YYYY'
            elif col_name in ('Numero', 'Prefixo', 'Instalação CCS', 'Situação', 'Modalidade'):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ('Endereço', 'Bairro'):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        limite = 60 if col_name == 'Endereço' else 35 if col_name == 'Bairro' else 45
        ws.column_dimensions[col_letter].width = min(max_len + 4, limite)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO(); wb.save(buffer); return buffer.getvalue()


def formatar_valor_corte(modalidade, valor):
    if valor is None or pd.isna(valor):
        return "—"
    if modalidade in ("ARRECADADO", "GRUPO A"):
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if modalidade == "QUANTIDADE DE FATURAS":
        return f"{int(valor)}"
    if modalidade == "SUSPENSÃO E VISTORIA P1":
        try: return pd.Timestamp(valor).strftime("%d/%m/%Y")
        except Exception: return str(valor)
    return str(valor)


# ══════════════════════════════════════════════════════════════════
# INTERFACE
# ══════════════════════════════════════════════════════════════════

st.markdown("""
<div class="main-header">
    <h1>⚡ Relatório de Serviços Prioritários</h1>
    <p>Automatize a geração do relatório diário com filtros inteligentes por modalidade.</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown('<p class="sidebar-title">📁 Entrada de Dados</p>', unsafe_allow_html=True)
    arquivo = st.file_uploader("Carregar arquivo", type=["xlsx", "xls", "xml"],
        help="Formatos aceitos: .xlsx, .xls, .xml (SpreadsheetML)")

if arquivo is None:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div style="text-align:center; padding: 3rem 1rem; color: #94a3b8;">
            <div style="font-size: 3rem; margin-bottom: 1rem;">📂</div>
            <div style="font-size: 1.1rem; font-weight: 600; color: #64748b;">Nenhum arquivo carregado</div>
            <div style="font-size: 0.9rem; margin-top: 0.5rem;">
                Use o painel lateral para carregar um arquivo <b>.xlsx</b>, <b>.xls</b> ou <b>.xml</b>.
            </div>
        </div>
        """, unsafe_allow_html=True)
    st.stop()

try:
    with st.spinner("📊 Lendo e processando o arquivo..."):
        df_raw = ler_arquivo(arquivo)
        df_clean = preparar_dados(df_raw)
except ValueError as e:
    st.error(f"❌ Erro ao processar o arquivo: {e}"); st.stop()
except Exception as e:
    st.error(f"❌ Erro inesperado: {e}"); st.stop()

total_raw = len(df_raw); total_clean = len(df_clean)
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown(f'<div class="metric-card"><div class="label">Total de Registros</div><div class="value">{total_raw:,}</div></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="metric-card"><div class="label">Registros "Atribuida"</div><div class="value">{total_clean:,}</div></div>', unsafe_allow_html=True)
with col3:
    perc = round((total_clean / total_raw * 100) if total_raw else 0, 1)
    st.markdown(f'<div class="metric-card"><div class="label">Taxa de Aproveitamento</div><div class="value">{perc}%</div></div>', unsafe_allow_html=True)
with col4:
    prefixos = df_clean['Prefixo'].nunique() if 'Prefixo' in df_clean.columns else 0
    st.markdown(f'<div class="metric-card"><div class="label">Prefixos Distintos</div><div class="value">{prefixos}</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

tab_relatorio, tab_mesclar, tab_cortes = st.tabs([
    "📄 Relatório", "🔀 Mesclar Modalidades", "✂️ Análise de Cortes"
])

# ─────────────────────────────────────────────────────────────
# ABA 1 — RELATÓRIO (modalidade única, comportamento original)
# ─────────────────────────────────────────────────────────────
with tab_relatorio:
    col_cfg, col_out = st.columns([1, 3])

    with col_cfg:
        st.markdown('<p class="sidebar-title">🎯 Configuração</p>', unsafe_allow_html=True)
        modalidade = st.selectbox("Modalidade de Prioridade", options=MODALIDADES,
            help="Selecione a lógica de filtragem e ordenação desejada.", key="rel_modalidade")
        st.markdown('<p class="sidebar-title">⚙️ Parâmetros</p>', unsafe_allow_html=True)
        limite_prefixo = st.number_input("Máx. de serviços por Prefixo", min_value=1, max_value=100, value=5, step=1, key="rel_limite_prefixo")
        limite_faturas = None; teto_faturas = None

        if modalidade == "ARRECADADO":
            st.markdown('<div class="info-box">🔹 Ordena por <b>Valor Faturas</b> (maior → menor). Exclui subtipos de Grupo A e Vistoria.</div>', unsafe_allow_html=True)
            limite_faturas = st.number_input("Limite máximo de Qtd. Faturas", min_value=1, max_value=99999, value=10, step=1, key="rel_limite_faturas")
        elif modalidade == "QUANTIDADE DE FATURAS":
            st.markdown('<div class="info-box">🔹 Ordena por <b>Quantidade de Faturas</b> (maior → menor).</div>', unsafe_allow_html=True)
            teto_faturas = st.number_input("Teto máximo de Qtd. Faturas", min_value=1, max_value=99999, value=10, step=1, key="rel_teto_faturas")
        elif modalidade == "SUSPENSÃO E VISTORIA P1":
            st.markdown('<div class="info-box">🔹 Filtra subtipos P1. Ordena por <b>Data Inclusão</b> (mais antiga) e desempata por <b>Valor Faturas</b> (maior).</div>', unsafe_allow_html=True)
        elif modalidade == "GRUPO A":
            st.markdown('<div class="info-box">🔹 Filtra subtipos Grupo A e ordena por <b>Valor Faturas</b> (maior → menor).</div>', unsafe_allow_html=True)

    with col_out:
        with st.spinner(f"⚙️ Aplicando filtros da modalidade **{modalidade}**..."):
            if modalidade == "ARRECADADO":
                df_result = modalidade_arrecadado(df_clean, limite_faturas, limite_prefixo)
            elif modalidade == "QUANTIDADE DE FATURAS":
                df_result = modalidade_quantidade_faturas(df_clean, teto_faturas, limite_prefixo)
            elif modalidade == "SUSPENSÃO E VISTORIA P1":
                df_result = modalidade_suspensao_p1(df_clean, limite_prefixo)
            elif modalidade == "GRUPO A":
                df_result = modalidade_grupo_a(df_clean, limite_prefixo)

        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:0.75rem;">
            <span class="badge">✅ {modalidade}</span>
            <span style="color:#64748b; font-size:0.9rem;">{len(df_result):,} registros encontrados</span>
        </div>
        """, unsafe_allow_html=True)

        if df_result.empty:
            st.warning("⚠️ Nenhum registro encontrado com os filtros aplicados. Tente ajustar os parâmetros.")
        else:
            st.dataframe(df_result, use_container_width=True, height=420, hide_index=True)
            st.markdown("<br>", unsafe_allow_html=True)
            excel_bytes = gerar_excel(df_result, modalidade)
            nome_arquivo = f"relatorio_{modalidade.lower().replace(' ', '_')}.xlsx"
            st.download_button(
                label=f"⬇️  Baixar Relatório Excel — {modalidade}",
                data=excel_bytes, file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="rel_download",
            )

# ─────────────────────────────────────────────────────────────
# ABA 2 — MESCLAR MODALIDADES
# Uma caixa de seleção por modalidade, com campo de quantidade
# logo abaixo. É necessário marcar pelo menos duas para mesclar.
# ─────────────────────────────────────────────────────────────
with tab_mesclar:
    st.markdown('<div class="info-box">🔹 Selecione ao menos <b>duas modalidades</b> e informe a quantidade de registros desejada para cada uma. Os registros de todas as modalidades marcadas serão combinados em um único relatório (sem duplicar o mesmo <b>Número</b>).</div>', unsafe_allow_html=True)

    cols_mesclar = st.columns(len(MODALIDADES))
    selecoes = {}
    for i, mod in enumerate(MODALIDADES):
        with cols_mesclar[i]:
            st.markdown(f'<div class="modal-card">', unsafe_allow_html=True)
            marcado = st.checkbox(mod, key=f"mesclar_check_{mod}")
            qtd = st.number_input("Quantidade", min_value=1, max_value=9999, value=10, step=1,
                                   key=f"mesclar_qtd_{mod}", disabled=not marcado)
            st.markdown('</div>', unsafe_allow_html=True)
            selecoes[mod] = {"marcado": marcado, "quantidade": qtd}

    modalidades_marcadas = [m for m, v in selecoes.items() if v["marcado"]]

    calcular = st.button("🔀 Mesclar Modalidades", key="mesclar_btn")

    if calcular:
        if len(modalidades_marcadas) < 2:
            st.error("⚠️ Selecione pelo menos duas modalidades para mesclar.")
        else:
            partes = []
            for mod in modalidades_marcadas:
                base = obter_base_modalidade(df_clean, mod)
                qtd = selecoes[mod]["quantidade"]
                parte = base.head(qtd).copy()
                parte = selecionar_colunas(parte)
                parte["Modalidade"] = mod
                partes.append(parte)

            df_mesclado = pd.concat(partes, ignore_index=True)
            antes_dedup = len(df_mesclado)
            if 'Numero' in df_mesclado.columns:
                df_mesclado = df_mesclado.drop_duplicates(subset='Numero', keep='first').reset_index(drop=True)
            duplicados_removidos = antes_dedup - len(df_mesclado)

            st.markdown(f"""
            <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:0.75rem;">
                <span class="badge">✅ {' + '.join(modalidades_marcadas)}</span>
                <span style="color:#64748b; font-size:0.9rem;">{len(df_mesclado):,} registros combinados
                {f' · {duplicados_removidos} duplicado(s) removido(s)' if duplicados_removidos else ''}</span>
            </div>
            """, unsafe_allow_html=True)

            if df_mesclado.empty:
                st.warning("⚠️ Nenhum registro encontrado com as modalidades e quantidades selecionadas.")
            else:
                st.dataframe(df_mesclado, use_container_width=True, height=420, hide_index=True)
                st.markdown("<br>", unsafe_allow_html=True)
                excel_bytes_mesclado = gerar_excel(df_mesclado, "MESCLADO")
                st.download_button(
                    label="⬇️  Baixar Relatório Mesclado",
                    data=excel_bytes_mesclado, file_name="relatorio_mesclado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="mesclar_download",
                )

# ─────────────────────────────────────────────────────────────
# ABA 3 — ANÁLISE DE CORTES
# O usuário define os critérios no topo e clica em calcular.
# Mostra, para cada modalidade, quantos registros resultam em
# cada limite de "serviços por Prefixo" e qual o valor de corte
# (o valor limite do último registro incluído). Apenas visual —
# não há botão de download nesta aba.
# ─────────────────────────────────────────────────────────────
with tab_cortes:
    st.markdown('<div class="info-box">🔹 Defina os critérios abaixo e clique em <b>Calcular Cortes</b> para visualizar, por modalidade, quantos registros resultam em cada limite de serviços por Prefixo e qual o valor de corte (o valor do último registro incluído no limite). Esta aba é apenas para análise — não gera arquivo para download.</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        modalidades_analise = st.multiselect("Modalidades a analisar", options=MODALIDADES, default=MODALIDADES, key="cortes_modalidades")
    with c2:
        limite_min = st.number_input("Limite mínimo por Prefixo", min_value=1, max_value=100, value=1, step=1, key="cortes_min")
    with c3:
        limite_max = st.number_input("Limite máximo por Prefixo", min_value=1, max_value=100, value=10, step=1, key="cortes_max")

    c4, c5 = st.columns(2)
    with c4:
        limite_faturas_cortes = st.number_input("Limite máx. de Qtd. Faturas (ARRECADADO)", min_value=1, max_value=99999, value=10, step=1, key="cortes_limite_faturas")
    with c5:
        teto_faturas_cortes = st.number_input("Teto máx. de Qtd. Faturas (QUANTIDADE DE FATURAS)", min_value=1, max_value=99999, value=10, step=1, key="cortes_teto_faturas")

    calcular_cortes = st.button("✂️ Calcular Cortes", key="cortes_btn")

    if calcular_cortes:
        if limite_min > limite_max:
            st.error("⚠️ O limite mínimo não pode ser maior que o limite máximo.")
        elif not modalidades_analise:
            st.error("⚠️ Selecione ao menos uma modalidade para analisar.")
        else:
            for mod in modalidades_analise:
                st.markdown(f"#### {mod}")
                kwargs = {}
                if mod == "ARRECADADO": kwargs["limite_faturas"] = limite_faturas_cortes
                if mod == "QUANTIDADE DE FATURAS": kwargs["teto_faturas"] = teto_faturas_cortes

                base = obter_base_modalidade(df_clean, mod, **kwargs)
                sort_col, _ = COLUNA_ORDENACAO[mod]

                linhas = []
                for limite in range(int(limite_min), int(limite_max) + 1):
                    grupo = aplicar_limite_prefixo(base, limite)
                    qtd_registros = len(grupo)
                    valor_corte = None
                    if qtd_registros > 0 and sort_col in grupo.columns:
                        if mod == "SUSPENSÃO E VISTORIA P1":
                            valor_corte = grupo[sort_col].max()
                        else:
                            valor_corte = grupo[sort_col].min()
                    linhas.append({
                        "Limite por Prefixo": limite,
                        "Qtd. Registros": qtd_registros,
                        "Valor de Corte": formatar_valor_corte(mod, valor_corte),
                    })

                df_cortes = pd.DataFrame(linhas)
                if base.empty:
                    st.warning("⚠️ Nenhum registro-base encontrado para esta modalidade com os critérios atuais.")
                else:
                    st.dataframe(df_cortes, use_container_width=True, hide_index=True)
                st.markdown("<br>", unsafe_allow_html=True)
